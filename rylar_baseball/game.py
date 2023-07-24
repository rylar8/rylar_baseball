from . import team, inning, player, atbat
import sqlite3
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as PYXLImage
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from PIL import Image as PILImage
import os
import matplotlib.pyplot as plt
from collections import Counter
from IPython.display import clear_output
import datetime

#Add a version or timestamp to the statistics. Check if there are trackman uploads connected to their ID with a newer timestamp.
#Only update those stats. 

class Game:
    def __init__(self):
        pass

    def loadCSV(self, csv, writeData = True):
        self.data = pd.read_csv(csv)
        self.stadium = self.data.iloc[0]['Stadium']
        self.league = self.data.iloc[0]['Level']
        self.division = self.data.iloc[0]['League']
        self.trackman_id = self.data.iloc[0]['GameID']
        self.date = pd.to_datetime(self.data.iloc[0]['UTCDateTime']).date()
        self.year = pd.to_datetime(self.data.iloc[0]['UTCDateTime']).year
        self.time = self.data.iloc[0]['Time']
        self.home = team.Team(self.data.iloc[0]['HomeTeam'])
        self.away = team.Team(self.data.iloc[0]['AwayTeam'])
        self.timestamp = int(datetime.datetime.now().timestamp())
        if writeData:
            self.toDatabase()
            self.updateStats()

    def loadDF(self, data, writeData = True):
        self.data = data
        self.stadium = self.data.iloc[0]['Stadium']
        self.league = self.data.iloc[0]['Level']
        self.division = self.data.iloc[0]['League']
        self.trackman_id = self.data.iloc[0]['GameID']
        self.date = pd.to_datetime(self.data.iloc[0]['UTCDateTime']).date()
        self.year = pd.to_datetime(self.data.iloc[0]['UTCDateTime']).year
        self.time = self.data.iloc[0]['Time']
        self.home = team.Team(self.data.iloc[0]['HomeTeam'])
        self.away = team.Team(self.data.iloc[0]['AwayTeam'])
        self.timestamp = int(datetime.datetime.now().timestamp())
        if writeData:
            self.toDatabase()
            self.updateStats()

    def loadID(self, game_id):
        conn = sqlite3.connect('rylar_baseball.db')
        cur = conn.cursor()

        cur.execute('''SELECT trackman.*, 
                    games.trackman_id AS trackman_game_id,
                    games.date AS Date, 
                    games.time AS Time,
                    batters.batter_name AS batter_name, batters.trackman_id AS batter_trackman_id, 
                    batter_side.side AS batter_side,
                    home_teams.team_name AS home_name, home_teams.trackman_name AS home_trackman_id,
                    away_teams.team_name AS away_name, away_teams.trackman_name AS away_trackman_id,
                    inning_split.split AS top_bottom,
                    pitchers.pitcher_name AS pitcher_name, pitchers.trackman_id AS pitcher_trackman_id,
                    pitcher_side.side AS pitcher_side,
                    catchers.catcher_name AS catcher_name, catchers.trackman_id AS catcher_trackman_id,
                    catcher_side.side AS catcher_side,
                    leagues.league_name AS league_name, leagues.trackman_name AS league_trackman_id,
                    divisions.division_name AS division_name, divisions.trackman_name AS division_trackman_id,
                    stadiums.stadium_name AS stadium_name, stadiums.trackman_name AS stadium_trackman_id,
                    auto_pitches.auto_pitch AS auto_pitch_type,
                    tagged_pitches.tagged_pitch AS tagged_pitch_type,
                    calls.call AS pitch_call,
                    hits.hit AS hit_type,
                    results.result AS result,
                    batter_teams.trackman_name AS batter_team_trackman_id,
                    pitcher_teams.trackman_name AS pitcher_team_trackman_id,
                    catcher_teams.trackman_name AS catcher_team_trackman_id,
                    k_or_bb.k_or_bb AS k_or_bb

                    FROM trackman 

                    LEFT JOIN games ON trackman.game_id = games.game_id
                    LEFT JOIN batters ON trackman.batter_id = batters.batter_id
                    LEFT JOIN sides AS batter_side ON batters.batter_side_id = batter_side.side_id
                    LEFT JOIN teams AS home_teams ON trackman.home_id = home_teams.team_id
                    LEFT JOIN teams AS away_teams ON trackman.away_id = away_teams.team_id
                    LEFT JOIN inning_split ON trackman.top_bottom_id = inning_split.split_id
                    LEFT JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
                    LEFT JOIN sides AS pitcher_side ON pitchers.pitcher_side_id = pitcher_side.side_id
                    LEFT JOIN catchers ON trackman.catcher_id = catchers.catcher_id
                    LEFT JOIN sides AS catcher_side ON catchers.catcher_side_id = catcher_side.side_id
                    LEFT JOIN leagues ON trackman.league_id = leagues.league_id
                    LEFT JOIN divisions ON trackman.division_id = divisions.division_id
                    LEFT JOIN stadiums ON games.stadium_id = stadiums.stadium_id
                    LEFT JOIN pitches AS auto_pitches ON trackman.auto_type_id = auto_pitches.type_id
                    LEFT JOIN pitches AS tagged_pitches ON trackman.tagged_type_id = tagged_pitches.type_id
                    LEFT JOIN calls ON trackman.call_id = calls.type_id
                    LEFT JOIN hits ON trackman.hit_type_id = hits.type_id
                    LEFT JOIN results ON trackman.result_id = results.type_id
                    LEFT JOIN teams AS batter_teams ON batters.team_id = batter_teams.team_id
                    LEFT JOIN teams AS pitcher_teams ON pitchers.team_id = pitcher_teams.team_id
                    LEFT JOIN teams AS catcher_teams ON catchers.team_id = catcher_teams.team_id
                    LEFT JOIN k_or_bb ON trackman.k_or_bb_id = k_or_bb.k_or_bb_id

                    WHERE games.trackman_id = ?''', (game_id,))
        #Set column names to match raw trackman columns
        cols = ['game_id', 'PitchNo', 'Inning', 'top_bottom_id', 'PAofInning', 'PitchofPA', 'pitcher_id', 'batter_id',
                'catcher_id', 'league_id', 'division_id', 'home_id', 'away_id', 'Outs', 'Balls', 'Strikes',
                'RelSpeed', 'VertBreak', 'InducedVertBreak', 'HorzBreak', 'SpinRate', 'SpinAxis', 'Tilt', 'RelHeight', 'RelSide', 
                'Extension', 'auto_type_id', 'tagged_type_id', 'call_id', 'PlateLocHeight', 'PlateLocSide', 'ExitSpeed',
                'Angle', 'Direction', 'HitSpinRate', 'hit_type_id', 'Distance', 'HangTime', 'Bearing', 'result_id',
                'OutsOnPlay', 'RunsScored', 'ThrowSpeed', 'PopTime', 'k_or_bb_id', 'VertApprAngle', 'HorzApprAngle', 'ZoneSpeed', 'ZoneTime', 'PositionAt110X', 'PositionAt110Y',
                'PositionAt110Z', 'LastTrackedDistance', 'pfxx', 'pfxz', 'x0', 'y0', 'z0', 'vx0', 'vy0', 'vz0', 'ax0', 'ay0', 'az0', 
                'ContactPositionX', 'ContactPositionY', 'ContactPositionZ', 'HitSpinAxis', 'GameID', 'Date', 'Time', 'Batter', 'BatterId',
                'BatterSide', 'home_name', 'HomeTeam', 'away_name', 'AwayTeam', 'Top/Bottom', 'Pitcher', 
                'PitcherId', 'PitcherThrows', 'Catcher', 'CatcherId', 'CatcherThrows', 'league_name', 
                'Level', 'division_name', 'League', 'stadium_name', 'Stadium', 'AutoPitchType',
                'TaggedPitchType', 'PitchCall', 'TaggedHitType', 'PlayResult', 'BatterTeam', 'PitcherTeam', 'CatcherTeam', 'KorBB']
        #Filter columns to get rid of database ids
        filt = ['GameID', 'Date', 'Time', 'PitchNo', 'Inning', 'Top/Bottom', 'PAofInning', 'PitchofPA', 'Pitcher',
                'PitcherId', 'PitcherThrows', 'PitcherTeam', 'Batter', 'BatterId', 'BatterSide', 'BatterTeam', 'Catcher',
                'CatcherId', 'CatcherThrows', 'CatcherTeam', 'league_name', 'Level', 'division_name', 'League',
                'home_name', 'HomeTeam', 'away_name', 'AwayTeam', 'Outs', 'Balls', 'Strikes', 'RelSpeed', 'VertBreak',
                'InducedVertBreak', 'HorzBreak', 'SpinRate', 'SpinAxis', 'Tilt', 'RelHeight', 'RelSide', 'Extension', 'AutoPitchType',
                'TaggedPitchType', 'PitchCall', 'PlateLocHeight','PlateLocSide', 'ExitSpeed', 'Angle', 'Direction',
                'HitSpinRate', 'TaggedHitType', 'Distance', 'HangTime', 'Bearing','PlayResult', 'KorBB', 'OutsOnPlay', 'RunsScored', 'ThrowSpeed',
                'PopTime', 'VertApprAngle', 'HorzApprAngle', 'ZoneSpeed', 'ZoneTime', 'PositionAt110X', 'PositionAt110Y',
                'PositionAt110Z', 'LastTrackedDistance', 'pfxx', 'pfxz', 'x0', 'y0', 'z0', 'vx0', 'vy0', 'vz0', 'ax0', 'ay0', 'az0', 
                'ContactPositionX', 'ContactPositionY', 'ContactPositionZ', 'HitSpinAxis', 'Stadium']
        
        self.data = pd.DataFrame(cur.fetchall(), columns=cols)[filt]
        self.stadium = self.data.iloc[0]['Stadium']
        self.league = self.data.iloc[0]['Level']
        self.division = self.data.iloc[0]['League']
        self.trackman_id = self.data.iloc[0]['GameID']
        self.date = pd.to_datetime(self.data.iloc[0]['Date']).date()
        self.year = pd.to_datetime(self.data.iloc[0]['Date']).year
        self.time = self.data.iloc[0]['Time']
        self.home = team.Team(self.data.iloc[0]['HomeTeam'])
        self.away = team.Team(self.data.iloc[0]['AwayTeam'])
        cur.execute('SELECT upload_timestamp FROM games WHERE trackman_id = ?', (game_id,))
        self.timestamp = cur.fetchone()[0]
        conn.close()

    def innings(self, top_bottom):
        innings = []
        for i in range(len(set(self.data[self.data['Top/Bottom'] == top_bottom.capitalize()]['Inning']))):
            innings.append(inning.Inning(self.data, i+1, top_bottom))
        return innings
    
    def batters(self):
        batter_ids = set(self.data['BatterId'])
        return [player.Batter(batter_id) for batter_id in batter_ids]

    def catchers(self):
        catcher_ids = set(self.data['CatcherId'])
        return [player.Catcher(catcher_id) for catcher_id in catcher_ids]

    def pitchers(self):
        pitcher_ids = set(self.data['PitcherId'])
        return [player.Pitcher(pitcher_id) for pitcher_id in pitcher_ids]

    def adjustZone(self, stadium_id):
        #Get a center of mass on the entire league and then adjust the points in this stadiums data so that the center of mass is the same
        pass
    
    def pitcherStatline(self, pitcher_id):
        #Define what a hit is
        hits = ['Single', 'Double', 'Triple', 'HomeRun']
        #Get pitcher data
        pitcher_data = self.data[self.data['PitcherId'] == pitcher_id]
        K = len(pitcher_data[pitcher_data['KorBB'] == 'Strikeout'])
        R = pitcher_data.RunsScored.sum()
        H = len(pitcher_data[pitcher_data['PlayResult'].isin(hits)])
        BB = len(pitcher_data[pitcher_data['KorBB'] == 'Walk'])
        HBP = len(pitcher_data[pitcher_data['PitchCall'] == 'HitByPitch'])

        return f'{K} K, {R} R, {H} H, {BB} BB, {HBP} HBP'

    def movementPlot(self, pitcher_id, view = 'pitcher'):
        if view == 'pitcher':
            mirror = 1
        elif view == 'catcher':
            mirror = -1
        else:
            raise Exception('View must be either "pitcher" or "catcher"')
        
        #Colors for pitch types
        pitch_colors = {'Fastball': '#FF0000', 'Four-Seam': '#FF0000', 'ChangeUp': '#00BFFF', 'Changeup': '#00BFFF', 'Slider': '#00FA9A',
                        'Cutter': '#7CFC00','Curveball': '#32CD32', 'Splitter': '#ADD8E6', 'Sinker': '#FF7F50', 'Knuckleball': '#48D1CC'}

        pitcher_data = self.data[self.data['PitcherId'] == pitcher_id]
        
        #Initialize plot
        fig, ax = plt.subplots(figsize = (2.5, 2.95))
        ax.set_xlim(-30,30)
        ax.set_ylim(-30,30)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['bottom'].set_visible(True)
        ax.spines['left'].set_visible(True)
        ax.spines['left'].set_position('zero')
        ax.spines['bottom'].set_position('zero')
        ax.axes.xaxis.set_visible(True)
        ax.axes.yaxis.set_visible(True)
        ax.spines['bottom'].set_linewidth(2)
        ax.spines['left'].set_linewidth(2)
        ax.spines['left'].set_color('black')
        ax.spines['bottom'].set_color('black')
        ax.tick_params(axis = 'both', which= 'both', bottom = False, left = False, labelleft = False, labelbottom = False)

        #Create lists for plotting
        x = []
        y = []
        c = []

        #Get pitches from tagged, if an error occurs use auto type
        try:
            pitches = set(pitcher_data['TaggedPitchType'].dropna())
            for pitch in pitches:
                pitch_data = pitcher_data[pitcher_data['TaggedPitchType'] == pitch]
                x.append(mirror * pitch_data['HorzBreak'].mean())
                y.append(pitch_data['InducedVertBreak'].mean())
                c.append(pitch_colors[pitch])
        except:
            pitches = set(pitcher_data['AutoPitchType'].dropna())
            for pitch in pitches:
                pitch_data = pitcher_data[pitcher_data['TaggedPitchType'] == pitch]
                x.append(mirror * pitch_data['HorzBreak'].mean())
                y.append(pitch_data['InducedVertBreak'].mean())
                c.append(pitch_colors[pitch])
        
        df = pd.DataFrame({'x': x,
                    'y': y,
                    'c': c})

        groups = df.groupby('c')
        
        for name, group in groups:
            plt.scatter(group.x, group.y, s = 100, zorder = 3, color = group.c, edgecolors= 'Black', linewidths = .45)
        
        #Save figure in temporary holding spot so it can be anchored in excel sheet
        plt.tight_layout()
        plt.savefig(f'temporary_figures//{self.date}{self.trackman_id}{pitcher_id}movement_plot.png', transparent = True)
        plt.close()

        return f'temporary_figures//{self.date}{self.trackman_id}{pitcher_id}movement_plot.png'
        
    def toDatabase(self):
        conn = sqlite3.connect('rylar_baseball.db')
        cur = conn.cursor()
        #Add league
        try:
            #Increment league id
            cur.execute('SELECT league_id FROM leagues')
            league_id = cur.fetchall()[-1][0] + 1
            #Add new league
            cur.execute('''INSERT INTO leagues (league_id, year, trackman_name)
            VALUES (?, ?, ?);''', (league_id, self.year, self.league))
            conn.commit()
            #Get full league name
            league_name = input(f"Input new league name ({self.league}): ")
            cur.execute('UPDATE leagues SET league_name = ? WHERE league_id = ?', (league_name, league_id))
            conn.commit()
        except:
            #Get league id
            cur.execute('SELECT league_id FROM leagues WHERE trackman_name = ? AND year = ?', (self.league, self.year))
            league_id = cur.fetchone()[0]
        #Add division
        try:
            #Increment division id
            cur.execute('SELECT division_id FROM divisions')
            division_id = cur.fetchall()[-1][0] + 1
            #Add new division
            cur.execute('''INSERT INTO divisions (division_id, league_id, year, trackman_name)
            VALUES (?, ?, ?, ?);''', (division_id, league_id, self.year, self.division))
            conn.commit()
            #Get full division name
            division_name = input(f"Input new division name ({self.division}): ")
            cur.execute('UPDATE divisions SET division_name = ? WHERE division_id = ?', (division_name, division_id))
            conn.commit()
        except:
            #Get division id
            cur.execute('SELECT division_id FROM divisions WHERE trackman_name = ? AND year = ?', (self.division, self.year))
            division_id = cur.fetchone()[0]
        #Add home team
        try:
            cur.execute('INSERT INTO teams (trackman_name, division_id, league_id, year) VALUES (?, ?, ?, ?)', 
                        (self.home.trackman_id, division_id, league_id, self.year))
            conn.commit()
            #Get full team name
            home_name = input(f"Input new team name ({self.home.trackman_id}): ")
            cur.execute('UPDATE teams SET team_name = ? WHERE trackman_name = ?', (home_name, self.home.trackman_id))
            conn.commit()
        except:
            try:
                #Update team division/league (when away team the league is not inputted)
                cur.execute('UPDATE teams SET division_id = ? AND league_id = ? WHERE trackman_name = ? AND year = ?', (self.home.trackman_id, self.year))
                conn.commit()
            except:
                pass
          
        #Add away team
        try:
            cur.execute('INSERT INTO teams (trackman_name, year) VALUES (?, ?)', 
                    (self.away.trackman_id, self.year))
            conn.commit()
            #Get full team name
            away_name = input(f"Input new team name ({self.away.trackman_id}): ")
            cur.execute('UPDATE teams SET team_name = ? WHERE trackman_name = ?', (away_name, self.away.trackman_id))
            conn.commit()
        except:
            pass
        #Add stadium
        try:
            cur.execute('INSERT INTO stadiums (trackman_name) VALUES (?)', (self.stadium,))
            conn.commit()
            #Get full stadium name
            stadium_name = input(f"Input new stadium name ({self.stadium}): ")
            cur.execute('UPDATE stadiums SET stadium_name = ? WHERE trackman_name = ?', (stadium_name, self.stadium))
            conn.commit()
        except:
            pass
        #Get stadium id
        cur.execute('SELECT stadium_id FROM stadiums WHERE trackman_name = ?', (self.stadium,))
        stadium_id = cur.fetchone()[0]
        #Get home id
        cur.execute('SELECT team_id FROM teams WHERE trackman_name = ? AND year = ?', (self.home.trackman_id, self.year))
        home_id = cur.fetchone()[0]
        #Get away id
        cur.execute('SELECT team_id FROM teams WHERE trackman_name = ? AND year = ?', (self.away.trackman_id, self.year))
        away_id = cur.fetchone()[0]
        try:
            #Add game
            cur.execute('''INSERT INTO games (trackman_id, date, time, stadium_id, league_id, division_id, home_id, away_id, upload_timestamp)
            VALUES (?,?,?,?,?,?,?,?,?)''', (self.trackman_id, str(self.date), self.time, stadium_id, league_id, division_id, home_id, away_id, self.timestamp))
            conn.commit()
        except:
            print(f'Game already in games table: {self.away.trackman_id} at {self.home.trackman_id} on {self.date} (trackman_id = {self.trackman_id})')
        #Add batters
        for batter_id in set(self.data['BatterId'].dropna()):
            batter_name = self.data[self.data['BatterId'] == batter_id].iloc[0]['Batter']
            #Get team id
            batter_team = self.data[self.data['BatterId'] == batter_id].iloc[0]['BatterTeam']
            cur.execute('SELECT team_id FROM teams WHERE trackman_name = ? AND year = ?',
                        (batter_team, self.year))
            team_id = cur.fetchone()[0]
            #Get batter side id
            batter_side = self.data[self.data['BatterId'] == batter_id].iloc[0]['BatterSide']
            cur.execute('SELECT side_id FROM sides WHERE side = ?',
                        (batter_side,))
            batter_side_id = cur.fetchone()[0]
            batter_id = int(batter_id)
            try:
                cur.execute('''INSERT INTO batters (batter_name, team_id, trackman_id, batter_side_id)
                VALUES (?, ?, ?, ?)''', (batter_name, team_id, batter_id, batter_side_id))
                conn.commit()
            except:
                pass
        #Add pitchers
        for pitcher_id in set(self.data['PitcherId'].dropna()):
            pitcher_name = self.data[self.data['PitcherId'] == pitcher_id].iloc[0]['Pitcher']
            #Get team id
            pitcher_team = self.data[self.data['PitcherId'] == pitcher_id].iloc[0]['PitcherTeam']
            cur.execute('SELECT team_id FROM teams WHERE trackman_name = ? AND year = ?',
                        (pitcher_team, self.year))
            team_id = cur.fetchone()[0]
            #Get pitcher side id
            pitcher_side = self.data[self.data['PitcherId'] == pitcher_id].iloc[0]['PitcherThrows']
            cur.execute('SELECT side_id FROM sides WHERE side = ?',
                        (pitcher_side,))
            pitcher_side_id = cur.fetchone()[0]
            pitcher_id = int(pitcher_id)
            try:
                cur.execute('''INSERT INTO pitchers (pitcher_name, team_id, trackman_id, pitcher_side_id)
                VALUES (?, ?, ?, ?)''', (pitcher_name, team_id, pitcher_id, pitcher_side_id))
                conn.commit()
            except:
                pass

        #Add catchers
        for catcher_id in set(self.data['CatcherId'].dropna()):
            catcher_name = self.data[self.data['CatcherId'] == catcher_id].iloc[0]['Catcher']
            #Get team id
            catcher_team = self.data[self.data['CatcherId'] == catcher_id].iloc[0]['PitcherTeam']
            cur.execute('SELECT team_id FROM teams WHERE trackman_name = ? AND year = ?',
                        (catcher_team, self.year))
            team_id = cur.fetchone()[0]
            #Get catcher side id
            catcher_side = self.data[self.data['CatcherId'] == catcher_id].iloc[0]['CatcherThrows']
            cur.execute('SELECT side_id FROM sides WHERE side = ?',
                        (catcher_side,))
            catcher_side_id = cur.fetchone()[0]
            catcher_id = int(catcher_id)
            try:
                cur.execute('''INSERT INTO catchers (catcher_name, team_id, trackman_id, catcher_side_id)
                VALUES (?, ?, ?, ?)''', (catcher_name, team_id, catcher_id, catcher_side_id))
                conn.commit()
            except:
                pass
        #Write Trackman data
        #Get game id
        cur.execute('SELECT game_id FROM games WHERE trackman_id = ?', (self.trackman_id,))
        game_id = cur.fetchone()[0]

        pitches = []
        for pitch in self.data.itertuples():
            pitch_num = pitch.PitchNo
            inning_ = pitch.Inning
            #Get top bottom id
            cur.execute('SELECT split_id FROM inning_split WHERE split = ?', (pitch._16,))
            top_bottom_id = cur.fetchone()[0]
            pa_of_inning = pitch.PAofInning
            pitch_of_pa = pitch.PitchofPA
            #Get pitcher id unless id does not exist, replace with example player
            try:
                cur.execute('SELECT pitcher_id FROM pitchers WHERE trackman_id = ?', (int(pitch.PitcherId),))
                pitcher_id = cur.fetchone()[0]
            except ValueError:
                pitcher_id = 0
            #Get batter id unless id does not exist, replace with example player
            try:
                cur.execute('SELECT batter_id FROM batters WHERE trackman_id = ?', (int(pitch.BatterId),))
                batter_id = cur.fetchone()[0]
            except ValueError:
                batter_id = 0
            #Get catcher id unless id does not exist, replace with example player
            try:
                cur.execute('SELECT catcher_id FROM catchers WHERE trackman_id = ?', (int(pitch.CatcherId),))
                catcher_id = cur.fetchone()[0]
            except ValueError:
                catcher_id = 0
            #Get league id
            cur.execute('SELECT league_id FROM leagues WHERE trackman_name = ? AND year = ?', (self.league, self.year))
            league_id = cur.fetchone()[0]
            #Get division id
            cur.execute('SELECT division_id FROM divisions WHERE trackman_name = ? AND year = ?', (self.division, self.year))
            division_id = cur.fetchone()[0]
            #Get home id
            cur.execute('SELECT team_id FROM teams WHERE trackman_name = ? AND year = ?', (self.home.trackman_id, self.year))
            home_id = cur.fetchone()[0]
            #Get away id
            cur.execute('SELECT team_id FROM teams WHERE trackman_name = ? AND year = ?', (self.away.trackman_id, self.year))
            away_id = cur.fetchone()[0]
            outs = pitch.Outs
            balls = pitch.Balls
            strikes = pitch.Strikes
            velocity = pitch.RelSpeed
            vertical = pitch.VertBreak
            induced = pitch.InducedVertBreak
            horizontal = pitch.HorzBreak
            spin = pitch.SpinRate
            axis = pitch.SpinAxis
            tilt = pitch.Tilt
            release_height = pitch.RelHeight
            release_side = pitch.RelSide
            release_extension = pitch.Extension
            #Get auto type id unless auto type is blank, set to null id
            try:
                cur.execute('SELECT type_id FROM pitches WHERE auto_pitch = ?', (pitch.AutoPitchType,))
                auto_type_id = cur.fetchone()[0]
            except TypeError:
                auto_type_id = 9
            #Get tagged type id unless tagged type is blank, set to null id
            try:
                cur.execute('SELECT type_id FROM pitches WHERE tagged_pitch = ?', (pitch.TaggedPitchType,))
                tagged_type_id = cur.fetchone()[0]
            except TypeError:
                tagged_type_id = 9
            #Get call id
            cur.execute('SELECT type_id FROM calls WHERE call = ?', (pitch.PitchCall,))
            try:
                call_id = cur.fetchone()[0]
            except:
                print(f'Unknown pitch call found in csv: "{pitch.PitchCall}" Pitch No. {pitch.number}\nPlease check database.')
            location_height = pitch.PlateLocHeight
            location_side = pitch.PlateLocSide
            exit_velocity = pitch.ExitSpeed
            launch_angle = pitch.Angle
            hit_direction = pitch.Direction
            hit_spin = pitch.HitSpinRate
            #Get hit type id
            cur.execute('SELECT type_id FROM hits WHERE hit = ?', (pitch.TaggedHitType,))
            try:
                hit_type_id = cur.fetchone()[0]
            except:
                print(f'Unknown hit type found in csv: "{pitch.TaggedHitType}" Pitch No. {pitch.number}\nPlease check database.')
            distance = pitch.Distance
            hang_time = pitch.HangTime
            hit_bearing = pitch.Bearing
            #Get result id
            cur.execute('SELECT type_id FROM results WHERE result = ?', (pitch.PlayResult,))
            try:
                result_id = cur.fetchone()[0]
            except:
                print(f'Unknown play result found in csv: "{pitch.PlayResult}" Pitch No. {pitch.number}\nPlease check database.')
            outs_made = pitch.OutsOnPlay
            runs_scored = pitch.RunsScored
            catcher_velocity = pitch.ThrowSpeed
            catcher_pop = pitch.PopTime
            #Get k_or_bb id
            cur.execute('SELECT k_or_bb_id FROM k_or_bb WHERE k_or_bb = ?', (pitch.KorBB,))
            try:
                k_or_bb_id = cur.fetchone()[0]
            except:
                print(f'Unknown K or BB value found in csv: "{pitch.KorBB}" Pitch No. {pitch.number}\nPlease check database.')
            vert_approach_angle = pitch.VertApprAngle
            horz_approach_angle = pitch.HorzApprAngle
            zone_speed = pitch.ZoneSpeed
            zone_time = pitch.ZoneTime
            pos_at_110x = pitch.PositionAt110X
            pos_at_110y = pitch.PositionAt110Y
            pos_at_110z = pitch.PositionAt110Z
            last_tracked_distance = pitch.LastTrackedDistance
            pfxx = pitch.pfxx
            pfxz = pitch.pfxz
            horz_loc_50 = pitch.x0
            from_home_loc_50 = pitch.y0
            vert_loc_50 = pitch.z0
            horz_velo_50 = pitch.vx0
            from_home_velo_50 = pitch.vy0
            vert_velo_50 = pitch.vz0
            horz_acc_50 = pitch.ax0
            from_home_acc_50 = pitch.ay0
            vert_acc_50 = pitch.az0
            con_pos_x = pitch.ContactPositionX
            con_pos_y = pitch.ContactPositionY
            con_pos_z = pitch.ContactPositionZ
            hit_spin_axis = pitch.HitSpinAxis

            pitches.append((game_id, pitch_num, inning_, top_bottom_id, pa_of_inning, pitch_of_pa,
            pitcher_id, batter_id, catcher_id, league_id, division_id, home_id, away_id, outs, balls, strikes, velocity, 
            vertical, induced, horizontal, spin, axis, tilt, release_height, release_side, release_extension, auto_type_id,
            tagged_type_id, call_id, location_height, location_side, exit_velocity, launch_angle, hit_direction, hit_spin, 
            hit_type_id, distance, hang_time, hit_bearing, result_id, outs_made, runs_scored, catcher_velocity, catcher_pop,
            k_or_bb_id, vert_approach_angle, horz_approach_angle, zone_speed, zone_time, pos_at_110x, pos_at_110y, pos_at_110z,
            last_tracked_distance, pfxx, pfxz, horz_loc_50, from_home_loc_50, vert_loc_50, horz_velo_50,
            from_home_velo_50, vert_velo_50, horz_acc_50, from_home_acc_50, vert_acc_50, con_pos_x, con_pos_y, con_pos_z, hit_spin_axis, self.timestamp))

        try:
            cur.executemany('''INSERT INTO trackman (game_id, pitch_num, inning, top_bottom_id, pa_of_inning, pitch_of_pa,
            pitcher_id, batter_id, catcher_id, league_id, division_id, home_id, away_id, outs, balls, strikes, velocity, 
            vertical, induced, horizontal, spin, axis, tilt, release_height, release_side, release_extension, auto_type_id,
            tagged_type_id, call_id, location_height, location_side, exit_velocity, launch_angle, hit_direction, hit_spin, 
            hit_type_id, distance, hang_time, hit_bearing, result_id, outs_made, runs_scored, catcher_velocity, catcher_pop, k_or_bb_id,
            vert_approach_angle, horz_approach_angle, zone_speed, zone_time, pos_at_110x, pos_at_110y, pos_at_110z, last_tracked_distance,
            pfxx, pfxz, horz_loc_50, home_loc_50, vert_loc_50,
            horz_velo_50, home_velo_50, vert_velo_50, horz_acc_50, home_acc_50, vert_acc_50, contact_pos_x,
            contact_pos_y, contact_pos_z, hit_spin_axis,upload_timestamp)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 
            ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,?)''', pitches)
            conn.commit()
        except:
            print(f'Data already in trackman table: {self.away.trackman_id} at {self.home.trackman_id} on {self.date} (game_id = {game_id}, trackman_id = {self.trackman_id})')
        conn.close()

    def writeBatterReports(self, team_id):
        temp_path = 'templates//postgame_batter_template.xlsx'
        #Get batters from data
        batters = set(self.data[self.data['BatterTeam'] == team_id]['BatterId'])

        #Define dictionaries to be used later
        pitches = {'Fastball' : 'FB' , 'Four-Seam': 'FB', 'ChangeUp' : 'CH', 'Changeup' : 'CH','Slider' : 'SL', 'Cutter' : 'CUT',
                'Curveball' : 'CB' , 'Splitter' : 'SP', 'Sinker' : '2FB', 'Knuckleball' : 'KN'}
        pitch_colors = {'Fastball': '00FF0000', 'Four-Seam': '00FF0000', 'ChangeUp': '0000BFFF', 'Changeup': '0000BFFF', 'Slider': '0000FA9A',
                'Cutter': '007CFC00','Curveball': '0032CD32', 'Splitter': '00ADD8E6', 'Sinker': '00FF7F50', 'Knuckleball': '0048D1CC'}
        hits = {'Popup' : 'Popup', 'LineDrive' : 'Line Drive', 'GroundBall' : 'Ground Ball', 'FlyBall' : 'Fly Ball', 'Bunt' : 'Bunt'}
        #Simplify pitch call to fit on sheet easier
        results = {'StrikeCalled' : 'Strike', 'StrikeSwinging' : 'Strike', 'FoulBall' : 'Foul', 'InPlay' : 'In Play',
                                'BallCalled' : 'Ball', 'HitByPitch' : 'HBP', 'BallinDirt' : 'Ball', 'Undefined' : '', 
                                'BallIntentional' : 'Ball'}
        #Use pitch call to decide if batter swung or not
        takes = ['StrikeCalled', 'BallCalled', 'HitByPitch', 'BallinDirt', 'BallIntentional']
        swings = ['InPlay', 'StrikeSwinging', 'FoulBall']

        #Write a new report for each batter id
        for batter_id in batters:
            #Initialize batter object
            batter = player.Batter(batter_id)
            #Initialize template
            wb = load_workbook(temp_path)
            ws = wb.active

            #Get a tuple of unique at bats for batter and sort in order
            at_bats = sorted(set(self.data[self.data['BatterId'] == batter_id][['PAofInning', 'Inning', 'Top/Bottom']].apply(lambda row : (row['Inning'], row['PAofInning'], row['Top/Bottom']), axis=1)), key= lambda tup: (tup[0], tup[1]))
        
            #Fill an at bat on the sheet for each at bat, use i to control what cell to write in
            i = 0
            for ab in at_bats:
                #Initialize at_bat object
                ws[f'J{i+10}'] = inning_ = ab[0]
                pa_of_inning = ab[1]
                top_bottom = ab[2].lower()
                at_bat = atbat.AtBat(self.data, inning_, top_bottom, pa_of_inning)
                
                ws[f'J{i+11}'] = at_bat.outs
                ws[f'J{i+12}'] = 'Coming Soon' #Home
                ws[f'J{i+13}'] = 'Coming Soon' #Away
                ws[f'J{i+14}'] = 'Coming Soon' #Runners

                #Initialize pitcher object
                pitcher = at_bat.pitcher()
                pitcher_id = pitcher.trackman_id

                ws[f'J{i+16}'] = pitcher.name.split(',')[0]
                ws[f'J{i+17}'] = pitcher.side

                ws['C3'] = ws['C53'] = batter.name #Player name
                ws['C5'] = ws['C55'] = self.date #Date
                ws['C7'] = ws['C57'] = f'v {pitcher.team_name.split()[-1]}' #Opponent

                #Try using the tagged fastball data, if an error occurs use the auto pitch type data
                #(embedded try/except statement because two different trackman versions exist)
                try:
                    pitcher_fastballs = self.data[((self.data['TaggedPitchType'] == 'Fastball') | (self.data['TaggedPitchType'] == 'Sinker')) & (self.data['PitcherId'] == pitcher_id)]
                except:
                    try:
                        pitcher_fastballs = self.data[((self.data['AutoPitchType'] == 'Fastball') | (self.data['AutoPitchType'] == 'Sinker')) & (self.data['PitcherId'] == pitcher_id)]
                    except:
                        pitcher_fastballs = self.data[((self.data['AutoPitchType'] == 'Four-Seam') | (self.data['AutoPitchType'] == 'Sinker')) & (self.data['PitcherId'] == pitcher_id)]
                #Get the mean and std pitcher fb velo
                mean_fb = round(pitcher_fastballs['RelSpeed'].dropna().mean())
                #If only 1 fastball is thrown then std throws and error, replace std with 0
                try:
                    std_fb = round(pitcher_fastballs['RelSpeed'].dropna().std())
                except:
                    std_fb = 0
                    ws[f'J{i+18}'] = f'{mean_fb-std_fb}-{mean_fb+std_fb} MPH'
                try:
                    #Get a list of the different pitches thrown, drop na
                    ws[f'J{i+19}'] = ','.join(set(self.data[self.data['PitcherId'] == pitcher_id]['TaggedPitchType'].dropna().map(pitches)))
                except:
                    ws[f'J{i+19}'] = ','.join(set(self.data[self.data['PitcherId'] == pitcher_id]['AutoPitchType'].dropna().map(pitches)))
                   
                #Try to get exit velo on last pitch of at bat, if an error occurs leave it blank
                try:
                    if not np.isnan(at_bat.pitches()[-1].exit_velocity):
                        ws[f'M{i+10}'] = f'{round(at_bat.pitches()[-1].exit_velocity, 2)} MPH'
                    else:
                        ws[f'M{i+10}'] = ''
                except:
                     ws[f'M{i+10}'] = ''
                #Try to get launch angle on last pitch of at bat, if an error occurs leave it blank
                try:
                     if not np.isnan(at_bat.pitches()[-1].launch_angle):
                        ws[f'M{i+12}']  = f'{round(at_bat.pitches()[-1].launch_angle, 2)}{chr(176)}'
                     else:
                        ws[f'M{i+12}'] = ''
                except:
                     ws[f'M{i+12}']  = ''
                #Try to get hit type on last pitch of at bat, if an error occurs leave it blank
                try:
                     ws[f'M{i+14}']  = hits[at_bat.pitches()[-1].hit_type]
                except:
                     ws[f'M{i+14}']  = ''
                #Try to get result on last pitch of at bat, if an error occurs leave it blank
                try:
                    if at_bat.pitches()[-1].result != 'Undefined':
                        ws[f'M{i+16}']  = at_bat.pitches()[-1].result
                    elif at_bat.pitches()[-1].k_or_bb != 'Undefined':
                        ws[f'M{i+16}']  = at_bat.pitches()[-1].k_or_bb
                    elif at_bat.pitches()[-1].call == 'HitByPitch':
                        ws[f'M{i+16}']  = 'Hit By Pitch'
                    else:
                        ws[f'M{i+16}'] = ''
                except:
                     ws[f'M{i+16}']  = ''
                ws[f'M{i+18}']  = 'Coming Soon' #QAB
                
                #Fill a pitch slot on the sheet for each pitch, use i and j to control what cell to write in
                #Do not exceed 8 pitches in the pitch/choice/result columns
                j = 0
                for pitch in at_bat.pitches()[:10]:
                    #Try tagged pitch type, if not use auto pitch type, if error leave blank
                    try:
                        ws[f'F{i+j+10}'] = pitches[pitch.tagged_type]
                        ws[f'F{i+j+10}'].fill = PatternFill('solid', fgColor=pitch_colors[pitch.tagged_type])
                    except:
                        try:
                            ws[f'F{i+j+10}'] = pitches[pitch.auto_type]
                            ws[f'F{i+j+10}'].fill = PatternFill('solid', fgColor=pitch_colors[pitch.auto_type])
                        except:
                            ws[f'F{i+j+10}'] = ''
                            ws[f'F{i+j+10}'].fill = PatternFill('solid', fgColor='00000000')
                    
                    if pitch.call in takes:
                        ws[f'G{i+j+10}'] = 'Take'
                    elif pitch.call in swings:
                        ws[f'G{i+j+10}'] = 'Swing'
                    else:
                        ws[f'G{i+j+10}'] = ''
        
                    ws[f'H{i+j+10}'] = results[pitch.call]
                    j+=1

                #Get image from at_bat object
                img_path = at_bat.zoneTracer(view='catcher')

                #Resize the image to better fit excel sheet using PIL library
                img = PILImage.open(img_path)
                img_resized = img.resize((int(img.width*.83), int(img.height*.55)))
                img_resized.save(img_path, 'PNG')
                img.close()

                #Add image to at_bat
                ws.add_image(PYXLImage(img_path), f'C{i+10}')

                #Jump to next at bat slot
                i += 11
                #Skip the second page header
                if i == 44:
                    i = 50
            #Create folders if they do not exist
            try:
                os.mkdir(f'postgame_batter_reports//{batter.team_trackman_id}')
            except:
                pass
            try:
                os.mkdir(f'postgame_batter_reports//{batter.team_trackman_id}//{self.date}')
            except:
                pass
            #Save file to folder with player name
            wb.save(f'postgame_batter_reports//{batter.team_trackman_id}//{self.date}//{batter.name}.xlsx')
            wb.close()
        
        #After writing all reports remove all of the temporary figures
        for filename in os.listdir('temporary_figures'):
            file_path = os.path.join('temporary_figures//', filename)
            os.remove(file_path)

    def writePitcherReports(self, team_id):
        conn = sqlite3.connect('rylar_baseball.db')
        cur = conn.cursor()

        temp_path = 'templates//postgame_pitcher_template.xlsx'
        #Get pitchers from data
        pitchers = set(self.data[self.data['PitcherTeam'] == team_id]['PitcherId'])

        #Define dictionaries to be used later
        pitches = {'Fastball' : 'FB' , 'Four-Seam': 'FB', 'ChangeUp' : 'CH', 'Changeup' : 'CH','Slider' : 'SL', 'Cutter' : 'CUT',
                'Curveball' : 'CB' , 'Splitter' : 'SP', 'Sinker' : '2FB', 'Knuckleball' : 'KN'}
        pitch_colors = {'FB': '00FF0000', 'FB': '00FF0000', 'CH': '0000BFFF', 'CH': '0000BFFF', 'SL': '0000FA9A',
                    'CUT': '007CFC00','CB': '0032CD32', 'SP': '00ADD8E6', '2FB': '00FF7F50', 'KN': '0048D1CC'}
        results = {'StrikeCalled' : 'Strike', 'StrikeSwinging' : 'Strike', 'FoulBall' : 'Foul', 'InPlay' : 'In Play',
                        'BallCalled' : 'Ball', 'HitByPitch' : 'HBP', 'BallinDirt' : 'Ball', 'Undefined' : '', 
                        'BallIntentional' : 'Ball'}

        #Write a new report for each pitcher id
        for pitcher_id in pitchers:
            #Initialize pitcher object
            pitcher = player.Pitcher(pitcher_id)
            #Initialize template
            wb = load_workbook(temp_path)
            ws = wb.active
            
            #Get pitcher's overall game data
            overall_data = self.data[self.data['PitcherId'] == pitcher_id]

            #Get opponent team name from database
            cur.execute('SELECT team_name FROM teams WHERE trackman_name = ?', (overall_data.BatterTeam.iloc[0],))
            opponent = cur.fetchone()[0]

            ws['C3'] = ws['C84'] = pitcher.name #Player name
            ws['C5'] = ws['C86'] = self.date #Date
            ws['C7'] = ws['C88'] = f'v {opponent.split()[-1]}' #Opponent
            
            #Try using the tagged pitch type data, if an error occurs use the auto pitch type data
            #(embedded try/except statement because two different trackman versions exist in league data)

            #Get a set of the pitch types
            try:
                ovr_pitch_types = set(overall_data['TaggedPitchType'].dropna())
            except:
                ovr_pitch_types = set(overall_data['AutoPitchType'].dropna())
            
            #To protect against there being no pitch tagging data, if there is only 1 tagged type (likely Undefined), then default to auto
            if len(ovr_pitch_types) <= 1:
                ovr_pitch_types = set(overall_data['AutoPitchType'].dropna())

            #Sort the pitches by usage
            pitch_counts = overall_data['TaggedPitchType'].value_counts()
            sorted_pitches = sorted(ovr_pitch_types, key=lambda pitch_type : pitch_counts.loc[pitch_type], reverse=True)
            
            #Fill a row for each pitch type, using i to control the row
            i = 0
            for pitch_type in sorted_pitches:
                #Get pitch type data
                try:
                    pitch_data = overall_data[overall_data['TaggedPitchType'] == pitch_type]
                except:
                    pitch_data = overall_data[overall_data['AutoPitchType'] == pitch_type]
    
                ws[f'F{i+10}'] = pitches[pitch_type] #Pitch type
                ws[f'F{i+10}'].fill = PatternFill('solid', fgColor=pitch_colors[pitches[pitch_type]])
                ws[f'G{i+10}'] = round(pitch_data['RelSpeed'].dropna().max(), 1) #Max velo
                ws[f'H{i+10}'] = round(pitch_data['RelSpeed'].dropna().mean(), 1) #Mean velo
                ws[f'I{i+10}'] = round(pitch_data['InducedVertBreak'].dropna().mean(), 1) #Vert break
                ws[f'J{i+10}'] = round(pitch_data['HorzBreak'].dropna().mean(), 1) #Horz break
                ws[f'K{i+10}'] = f'{round(len(pitch_data) / len(overall_data) * 100, 1)}%' #Usage
        
                #Get strike rate
                calls = pitch_data['PitchCall'].map(results)
                ws[f'L{i+10}'] = f'{round(len(calls[(calls == "Strike") | (calls == "Foul") | (calls == "In Play")]) / len(pitch_data) * 100, 1)}%'
                
                #Get hard hit rate, leave blank if no balls in play
                if len(pitch_data[pitch_data["PitchCall"] == "InPlay"]) == 0:
                    ws[f'M{i+10}'] = '-'
                    ws[f'M{i+10}'].font = Font(bold=True, italic=True, size=20)
                else:
                    ws[f'M{i+10}'] = f'{round(len(pitch_data[(pitch_data["ExitSpeed"] >= 95) & (pitch_data["PitchCall"] == "InPlay")]) / len(pitch_data[pitch_data["PitchCall"] == "InPlay"]) * 100, 1)}%'

                #Get chase rate, leave blank if no balls thrown
                #Get balls based on universal strike zone
                balls = pitch_data[(pitch_data['PlateLocSide'] < -0.7508) | (pitch_data['PlateLocSide'] > 0.7508) | (pitch_data['PlateLocHeight'] < 1.5942) | (pitch_data['PlateLocHeight'] > 3.6033)]
                chase = balls[(balls['PitchCall'] == 'FoulBall') | (balls['PitchCall'] == 'InPlay') | (balls['PitchCall'] == 'StrikeSwinging')]
                if len(balls) == 0:
                    ws[f'N{i+10}'] = '-'
                    ws[f'N{i+10}'].font = Font(bold=True, italic=True, size=20)
                else:
                    ws[f'N{i+10}'] = f'{round(len(chase) / len(balls) * 100, 1)}%'

                #Get whiff rate, leave blank if no swings
                swings = pitch_data[(pitch_data['PitchCall'] == 'FoulBall') | (pitch_data['PitchCall'] == 'InPlay') | (pitch_data['PitchCall'] == 'StrikeSwinging')]
                whiffs = swings[swings['PitchCall'] == 'StrikeSwinging']
                if len(swings) == 0:
                    ws[f'O{i+10}'] = '-'
                    ws[f'O{i+10}'].font = Font(bold=True, italic=True, size=20)
                else:
                    ws[f'O{i+10}'] = f'{round(len(whiffs) / len(swings) * 100, 1)}%'
                
                #Advance to next pitch, if pitcher threw 6+ pitch types then skip the least used pitches
                i += 2
                if i == 10:
                    break

            #Fill overall data

            #Get strike rate
            calls = overall_data['PitchCall'].map(results)
            ws[f'L{20}'] = f'{round(len(calls[(calls == "Strike") | (calls == "Foul") | (calls == "In Play")]) / len(overall_data) * 100, 1)}%'
            
            #Get hard hit rate, leave blank if no balls in play
            if len(overall_data[overall_data["PitchCall"] == "InPlay"]) == 0:
                ws[f'M{20}'] = '-'
                ws[f'M{20}'].font = Font(bold=True, italic=True, size=20)
            else:
                ws[f'M{20}'] = f'{round(len(overall_data[(overall_data["ExitSpeed"] >= 95) & (overall_data["PitchCall"] == "InPlay")]) / len(overall_data[overall_data["PitchCall"] == "InPlay"]) * 100, 1)}%'

            #Get chase rate, leave blank if no balls thrown
            #Get balls based on universal strike zone
            balls = overall_data[(overall_data['PlateLocSide'] < -0.7508) | (overall_data['PlateLocSide'] > 0.7508) | (overall_data['PlateLocHeight'] < 1.5942) | (overall_data['PlateLocHeight'] > 3.6033)]
            chase = balls[(balls['PitchCall'] == 'FoulBall') | (balls['PitchCall'] == 'InPlay') | (balls['PitchCall'] == 'StrikeSwinging')]
            if len(balls) == 0:
                ws[f'N{20}'] = '-'
                ws[f'N{20}'].font = Font(bold=True, italic=True, size=20)
            else:
                ws[f'N{20}'] = f'{round(len(chase) / len(balls) * 100, 1)}%'

            #Get whiff rate, leave blank if no swings
            swings = overall_data[(overall_data['PitchCall'] == 'FoulBall') | (overall_data['PitchCall'] == 'InPlay') | (overall_data['PitchCall'] == 'StrikeSwinging')]
            whiffs = swings[swings['PitchCall'] == 'StrikeSwinging']
            if len(swings) == 0:
                ws[f'O{20}'] = '-'
                ws[f'O{20}'].font = Font(bold=True, italic=True, size=20)
            else:
                ws[f'O{20}'] = f'{round(len(whiffs) / len(swings) * 100, 1)}%'

            ws[f'G{22}'] = self.pitcherStatline(pitcher_id)

            #Get image from game method
            img_path = self.movementPlot(pitcher_id)

            #Resize the image to better fit excel sheet using PIL library
            img = PILImage.open(img_path)
            img_resized = img.resize((int(img.width*.83), int(img.height*.96)))
            img_resized.save(img_path, 'PNG')
            img.close()

            #Add image to overall
            ws.add_image(PYXLImage(img_path), f'C{10}')

            #Get a tuple of unique innings for pitcher and sort in order
            innings = sorted(set(self.data[self.data['PitcherId'] == pitcher_id][['Inning', 'Top/Bottom']].apply(lambda row : (row['Inning'], row['Top/Bottom']), axis=1)), key= lambda tup: (tup[0], tup[1]))

            #Fill an inning on the sheet for each inning pitched, use j to control what cell to write in
            j = 15
            for inn in innings:
                #Initialize inning object
                inn_num = inn[0]
                top_bottom = inn[1].lower()
                inning_ = inning.Inning(self.data, inn_num, top_bottom)

                #Get pitcher's inning data
                inning_data = inning_.data[inning_.data['PitcherId'] == pitcher_id]
                
                #Try using the tagged pitch type data, if an error occurs use the auto pitch type data
                #(embedded try/except statement because two different trackman versions exist in league data)

                #Get a set of the pitch types
                try:
                    inn_pitch_types = set(inning_data['TaggedPitchType'].dropna())
                except:
                    inn_pitch_types = set(inning_data['AutoPitchType'].dropna())
                
                #To protect against there being no pitch tagging data, if there is only 1 tagged type (likely Undefined), then default to auto
                if len(inn_pitch_types) <= 1:
                    inn_pitch_types = set(inning_data['AutoPitchType'].dropna())

                #Sort the pitches by usage
                pitch_counts = inning_data['TaggedPitchType'].value_counts()
                sorted_pitches = sorted(inn_pitch_types, key=lambda pitch_type : pitch_counts.loc[pitch_type], reverse=True)

                #Fill a row for each pitch type, using i to control the row
                i = 0
                for pitch_type in sorted_pitches:
                    #Get pitch type data
                    try:
                        pitch_data = inning_data[inning_data['TaggedPitchType'] == pitch_type]
                    except:
                        pitch_data = inning_data[inning_data['AutoPitchType'] == pitch_type]
        
                    ws[f'F{i+10+j}'] = pitches[pitch_type] #Pitch type
                    ws[f'F{i+10+j}'].fill = PatternFill('solid', fgColor=pitch_colors[pitches[pitch_type]])
                    ws[f'G{i+10+j}'] = round(pitch_data['RelSpeed'].dropna().max(), 1) #Max velo
                    ws[f'H{i+10+j}'] = round(pitch_data['RelSpeed'].dropna().mean(), 1) #Mean velo
                    ws[f'I{i+10+j}'] = round(pitch_data['InducedVertBreak'].dropna().mean(), 1) #Vert break
                    ws[f'J{i+10+j}'] = round(pitch_data['HorzBreak'].dropna().mean(), 1) #Horz break
                    ws[f'K{i+10+j}'] = f'{round(len(pitch_data) / len(overall_data) * 100, 1)}%' #Usage
            
                    #Get strike rate
                    calls = pitch_data['PitchCall'].map(results)
                    ws[f'L{i+10+j}'] = f'{round(len(calls[(calls == "Strike") | (calls == "Foul") | (calls == "In Play")]) / len(pitch_data) * 100, 1)}%'
                    
                    #Get hard hit rate, leave blank if no balls in play
                    if len(pitch_data[pitch_data["PitchCall"] == "InPlay"]) == 0:
                        ws[f'M{i+10+j}'] = '-'
                        ws[f'M{i+10+j}'].font = Font(bold=True, italic=True, size=20)
                    else:
                        ws[f'M{i+10+j}'] = f'{round(len(pitch_data[(pitch_data["ExitSpeed"] >= 95) & (pitch_data["PitchCall"] == "InPlay")]) / len(pitch_data[pitch_data["PitchCall"] == "InPlay"]) * 100, 1)}%'

                    #Get chase rate, leave blank if no balls thrown
                    #Get balls based on universal strike zone
                    balls = pitch_data[(pitch_data['PlateLocSide'] < -0.7508) | (pitch_data['PlateLocSide'] > 0.7508) | (pitch_data['PlateLocHeight'] < 1.5942) | (pitch_data['PlateLocHeight'] > 3.6033)]
                    chase = balls[(balls['PitchCall'] == 'FoulBall') | (balls['PitchCall'] == 'InPlay') | (balls['PitchCall'] == 'StrikeSwinging')]
                    if len(balls) == 0:
                        ws[f'N{i+10+j}'] = '-'
                        ws[f'N{i+10+j}'].font = Font(bold=True, italic=True, size=20)
                    else:
                        ws[f'N{i+10+j}'] = f'{round(len(chase) / len(balls) * 100, 1)}%'

                    #Get whiff rate, leave blank if no swings
                    swings = pitch_data[(pitch_data['PitchCall'] == 'FoulBall') | (pitch_data['PitchCall'] == 'InPlay') | (pitch_data['PitchCall'] == 'StrikeSwinging')]
                    whiffs = swings[swings['PitchCall'] == 'StrikeSwinging']
                    if len(swings) == 0:
                        ws[f'O{i+10+j}'] = '-'
                        ws[f'O{i+10+j}'].font = Font(bold=True, italic=True, size=20)
                    else:
                        ws[f'O{i+10+j}'] = f'{round(len(whiffs) / len(swings) * 100, 1)}%'
                    
                    #Advance to next pitch, if pitcher threw 6+ pitch types then skip the least used pitches
                    i += 2
                    if i == 10:
                        break

                #Fill overall data

                #Get strike rate
                calls = inning_data['PitchCall'].map(results)
                ws[f'L{20+j}'] = f'{round(len(calls[(calls == "Strike") | (calls == "Foul") | (calls == "In Play")]) / len(inning_data) * 100, 1)}%'
                
                #Get hard hit rate, leave blank if no balls in play
                if len(inning_data[inning_data["PitchCall"] == "InPlay"]) == 0:
                    ws[f'M{20+j}'] = '-'
                    ws[f'M{20+j}'].font = Font(bold=True, italic=True, size=20)
                else:
                    ws[f'M{20+j}'] = f'{round(len(inning_data[(inning_data["ExitSpeed"] >= 95) & (inning_data["PitchCall"] == "InPlay")]) / len(inning_data[inning_data["PitchCall"] == "InPlay"]) * 100, 1)}%'

                #Get chase rate, leave blank if no balls thrown
                #Get balls based on universal strike zone
                balls = inning_data[(inning_data['PlateLocSide'] < -0.7508) | (inning_data['PlateLocSide'] > 0.7508) | (inning_data['PlateLocHeight'] < 1.5942) | (inning_data['PlateLocHeight'] > 3.6033)]
                chase = balls[(balls['PitchCall'] == 'FoulBall') | (balls['PitchCall'] == 'InPlay') | (balls['PitchCall'] == 'StrikeSwinging')]
                if len(balls) == 0:
                    ws[f'N{20+j}'] = '-'
                    ws[f'N{20+j}'].font = Font(bold=True, italic=True, size=20)
                else:
                    ws[f'N{20+j}'] = f'{round(len(chase) / len(balls) * 100, 1)}%'

                #Get whiff rate, leave blank if no swings
                swings = inning_data[(inning_data['PitchCall'] == 'FoulBall') | (inning_data['PitchCall'] == 'InPlay') | (inning_data['PitchCall'] == 'StrikeSwinging')]
                whiffs = swings[swings['PitchCall'] == 'StrikeSwinging']
                if len(swings) == 0:
                    ws[f'O{20+j}'] = '-'
                    ws[f'O{20+j}'].font = Font(bold=True, italic=True, size=20)
                else:
                    ws[f'O{20+j}'] = f'{round(len(whiffs) / len(swings) * 100, 1)}%'

                ws[f'G{22+j}'] = inning_.pitcherStatline(pitcher_id)

                #Get image from inning method
                img_path = inning_.movementPlot(pitcher_id)

                #Resize the image to better fit excel sheet using PIL library
                img = PILImage.open(img_path)
                img_resized = img.resize((int(img.width*.83), int(img.height*.96)))
                img_resized.save(img_path, 'PNG')
                img.close()

                #Add image to overall
                ws.add_image(PYXLImage(img_path), f'C{10+j}')

                j += 15

                if j == 75:
                    j = 81

             #Create folders if they do not exist
            try:
                os.mkdir(f'postgame_pitcher_reports//{pitcher.team_trackman_id}')
            except:
                pass
            try:
                os.mkdir(f'postgame_pitcher_reports//{pitcher.team_trackman_id}//{self.date}')
            except:
                pass
            #Save file to folder with player name
            wb.save(f'postgame_pitcher_reports//{pitcher.team_trackman_id}//{self.date}//{pitcher.name}.xlsx')
            wb.close()
        
        #After writing all reports remove all of the temporary figures
        for filename in os.listdir('temporary_figures'):
            file_path = os.path.join('temporary_figures//', filename)
            os.remove(file_path)

        conn.close()

    def setupDatabase(self):

        conn = sqlite3.connect('rylar_baseball.db')
        cur = conn.cursor()
        
        cur.executescript('''
        CREATE PROCEDURE update_batting(b_id INTEGER, ts INTEGER, l_id INTEGER, d_id INTEGER, t_id INTEGER, yr INTEGER) AS
        BEGIN
            DELETE FROM batting_stats_standard WHERE batter_id = b_id;
            INSERT INTO batting_stats_standard (batter_id, league_id, division_id, team_id, year, last_updated) VALUES (b_id, l_id, d_id, t_id, yr, ts);
            WITH 
                batter_g_results AS (
                    SELECT batter_id, COUNT(DISTINCT game_id) AS g_count
                    FROM trackman
                    WHERE (upload_timestamp + 43200) > ts
                    AND batter_id = b_id
                    GROUP BY batter_id
                ),
                batter_ab_results AS (
                    SELECT batter_id, COUNT(*) AS ab_count
                    FROM (
                        SELECT *, ROW_NUMBER() OVER (PARTITION BY batter_id, game_id, inning, pa_of_inning ORDER BY pitch_num DESC) AS row_num
                        FROM trackman
                        WHERE (upload_timestamp + 43200) > ts
                        AND batter_id = b_id
                    ) t
                    WHERE (row_num = 1) AND (k_or_bb_id != 3) AND (result_id != 7) AND (call_id != 6)
                    GROUP BY batter_id
                ),
                batter_pa_results AS (
                    SELECT batter_id, COUNT(DISTINCT (game_id || inning || pa_of_inning)) AS pa_count
                    FROM trackman
                    WHERE (upload_timestamp + 43200) > ts
                    AND batter_id = b_id
                    GROUP BY batter_id
                ),
                batter_h_results AS (
                    SELECT batter_id, COUNT(*) AS h_count
                    FROM trackman
                    WHERE result_id <= 4 AND (upload_timestamp + 43200) > ts
                    AND batter_id = b_id
                    GROUP BY batter_id
                ),
                batter_2b_results AS (
                    SELECT batter_id, COUNT(*) AS _2b_count
                    FROM trackman
                    WHERE result_id = 2 AND (upload_timestamp + 43200) > ts
                    AND batter_id = b_id
                    GROUP BY batter_id
                ),
                batter_3b_results AS (
                    SELECT batter_id, COUNT(*) AS _3b_count
                    FROM trackman
                    WHERE result_id = 3 AND (upload_timestamp + 43200) > ts
                    AND batter_id = b_id
                    GROUP BY batter_id
                ),
                batter_hr_results AS (
                    SELECT batter_id, COUNT(*) AS hr_count
                    FROM trackman
                    WHERE result_id = 4 AND (upload_timestamp + 43200) > ts
                    AND batter_id = b_id
                    GROUP BY batter_id
                ),
                batter_rbi_results AS (
                    SELECT batter_id, SUM(runs_scored) AS rbi_count
                    FROM trackman
                    WHERE runs_scored >= 1 AND outs_made < 2 AND (upload_timestamp + 43200) > ts
                    AND batter_id = b_id
                    GROUP BY batter_id
                ),
                batter_bb_results AS (
                    SELECT batter_id, COUNT(*) AS bb_count
                    FROM trackman
                    WHERE k_or_bb_id = 3 AND call_id != 9 AND (upload_timestamp + 43200) > ts
                    AND batter_id = b_id
                    GROUP BY batter_id
                ),
                batter_ibb_results AS (
                    SELECT batter_id, COUNT(*) AS ibb_count
                    FROM trackman
                    WHERE k_or_bb_id = 3 AND call_id = 9 AND (upload_timestamp + 43200) > ts
                    AND batter_id = b_id
                    GROUP BY batter_id
                ),
                batter_so_results AS (
                    SELECT batter_id, COUNT(*) AS so_count
                    FROM trackman
                    WHERE k_or_bb_id = 2 AND (upload_timestamp + 43200) > ts
                    AND batter_id = b_id
                    GROUP BY batter_id
                ),
                batter_hbp_results AS (
                    SELECT batter_id, COUNT(*) AS hbp_count
                    FROM trackman
                    WHERE call_id = 6 AND (upload_timestamp + 43200) > ts
                    AND batter_id = b_id
                    GROUP BY batter_id
                ),
                batter_sac_results AS (
                    SELECT batter_id, COUNT(*) AS sac_count
                    FROM trackman
                    WHERE result_id = 7 AND (upload_timestamp + 43200) > ts
                    AND batter_id = b_id
                    GROUP BY batter_id
                ),
                batter_gdp_results AS (
                    SELECT batter_id, COUNT(*) AS gdp_count
                    FROM trackman
                    WHERE outs_made >= 2 AND hit_type_id = 4 AND (upload_timestamp + 43200) > ts
                    AND batter_id = b_id
                    GROUP BY batter_id
                ),
                batter_tb_results AS (
                    SELECT batter_id, (h_count - _2b_count - _3b_count - hr_count + 2*_2b_count + 3*_3b_count + 4*hr_count) AS tb_count
                    FROM batter_h_results
                    LEFT JOIN batter_2b_results USING (batter_id)
                    LEFT JOIN batter_3b_results USING (batter_id)
                    LEFT JOIN batter_hr_results USING (batter_id)
                    WHERE batter_id = b_id
                )
                          
            UPDATE batting_stats_standard
            SET g = (SELECT g_count FROM batter_g_results),
                ab = (SELECT ab_count FROM batter_ab_results),
                pa = (SELECT pa_count FROM batter_pa_results),
                h = (SELECT h_count FROM batter_h_results),
                _2b = (SELECT _2b_count FROM batter_2b_results),
                _3b = (SELECT _3b_count FROM batter_3b_results),
                hr = (SELECT hr_count FROM batter_hr_results),
                rbi = (SELECT rbi_count FROM batter_rbi_results),
                bb = (SELECT bb_count FROM batter_bb_results),
                ibb = (SELECT ibb_count FROM batter_ibb_results),
                so = (SELECT so_count FROM batter_so_results),
                hbp = (SELECT hbp_count FROM batter_hbp_results),
                sac = (SELECT sac_count FROM batter_sac_results),
                gdp = (SELECT gdp_count FROM batter_gdp_results),
                tb = (SELECT tb_count FROM batter_tb_results)
            WHERE batter_id = b_id;
            
            DELETE FROM batting_stats_statcast WHERE batter_id = b_id;
            INSERT INTO batting_stats_statcast (batter_id, league_id, division_id, team_id, year, last_updated) VALUES (b_id, l_id, d_id, t_id, yr, ts);

            WITH 
                batter_bbe_results AS (
                    SELECT batter_id, COUNT(*) AS bbe_count
                    FROM trackman
                    WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ts
                    AND batter_id = b_id
                    GROUP BY batter_id
                ),
                batter_avg_ev_results AS (
                    SELECT batter_id, AVG(exit_velocity) AS avg_ev_value
                    FROM trackman
                    WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ts
                    AND batter_id = b_id
                    GROUP BY batter_id
                ),
                batter_max_ev_results AS (
                    SELECT batter_id, MAX(exit_velocity) AS max_ev_value
                    FROM trackman
                    WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ts
                    AND batter_id = b_id
                    GROUP BY batter_id
                ),
                batter_avg_la_results AS (
                    SELECT batter_id, AVG(launch_angle) AS avg_la_value
                    FROM trackman
                    WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ts
                    AND batter_id = b_id
                    GROUP BY batter_id
                ),

                batter_brl_results AS (
                    SELECT batter_id, SUM(barrel_count) AS brls_count
                    FROM (
                        SELECT 
                            batter_id, 
                            CASE 
                                WHEN exit_velocity >= 98.0 AND exit_velocity <= 99.0 AND launch_angle >= 26.0 AND launch_angle <= 30.0 THEN 1
                                WHEN exit_velocity >= 98.0 AND exit_velocity <= 100.0 AND launch_angle >= 25.0 AND launch_angle <= 31.0 THEN 1
                                WHEN exit_velocity > 100.0 THEN
                                    CASE
                                        WHEN launch_angle >= (MAX(25.0 - ((exit_velocity - 100.0) * 1.2), 8.0)) THEN 1
                                        WHEN launch_angle <= (MIN(31.0 + ((exit_velocity - 100.0) * 1.2), 50.0)) THEN 1
                                        ELSE 0
                                    END
                                ELSE 0
                            END AS barrel_count
                        FROM trackman
                        WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ts AND batter_id = b_id
                    ) AS barrels
                    GROUP BY batter_id
                ),
                          
                batter_hh_results AS (
                    SELECT batter_id, COUNT(*) AS hh_count
                    FROM trackman
                    WHERE call_id = 4 AND exit_velocity >= 95 AND (upload_timestamp + 43200) > ts
                    AND batter_id = b_id
                    GROUP BY batter_id
                ),
                batter_hh_rate_results AS (
                    SELECT batter_id, hh_count / bbe_count AS hh_rate_value
                    FROM (
                        SELECT 
                            s.batter_id, 
                            s.bbe AS bbe_count,
                            h.hh_count
                        FROM batting_stats_statcast s
                        LEFT JOIN batter_hh_results h ON s.batter_id = h.batter_id
                        WHERE (s.last_updated + 43200) > ts
                    )
                )
                          
            UPDATE batting_stats_statcast
            SET bbe = (SELECT bbe_count FROM batter_bbe_results WHERE batter_id = b_id),
                avg_ev = (SELECT avg_ev_value FROM batter_avg_ev_results WHERE batter_id = b_id),
                max_ev = (SELECT max_ev_value FROM batter_max_ev_results WHERE batter_id = b_id),
                avg_la = (SELECT avg_la_value FROM batter_avg_la_results WHERE batter_id = b_id),
                brls = (SELECT brls_count FROM batter_barrel_results WHERE batter_id = b_id),
                hh = (SELECT hh_count FROM batter_hh_results WHERE batter_id = b_id),
                hh_rate = (SELECT hh_rate_value FROM batter_hh_rate_results WHERE batter_id = b_id)
            WHERE batter_id = b_id;
                          
            DELETE FROM batting_stats_advanced WHERE batter_id = b_id;
            INSERT INTO batting_stats_advanced (batter_id, league_id, division_id, team_id, year, last_updated) VALUES (b_id, l_id, d_id, t_id, yr, ts);

            WITH 
                batter_avg_results AS (
                    SELECT batter_id, ab, h
                    FROM batting_stats_standard
                    WHERE (last_updated + 43200) > ts
                ),
                batter_obp_results AS (
                    SELECT batter_id, pa, h, bb, hbp
                    FROM batting_stats_standard
                    WHERE (last_updated + 43200) > ts
                ),
                batter_slg_results AS (
                    SELECT batter_id, ab, tb
                    FROM batting_stats_standard
                    WHERE (last_updated + 43200) > ts
                ),
                batter_ops_results AS (
                    SELECT batter_id, slg, obp
                    FROM batting_stats_advanced
                    WHERE (last_updated + 43200) > ts
                ),
                batter_iso_results AS (
                    SELECT batter_id, avg, slg
                    FROM batting_stats_advanced
                    WHERE (last_updated + 43200) > ts
                ),
                batter_babip_results AS (
                    SELECT batter_id, COUNT(*) AS batted_ball_events, COUNT(CASE WHEN result_id <= 4 THEN 1 END) AS hits
                    FROM trackman
                    WHERE call_id = 4 AND (upload_timestamp + 43200) > ts
                    GROUP BY batter_id
                ),
                batter_bb_rate_results AS (
                    SELECT batter_id, pa, bb
                    FROM batting_stats_standard
                    WHERE (last_updated + 43200) > ts
                ),
                batter_so_rate_results AS (
                    SELECT batter_id, pa, so
                    FROM batting_stats_standard
                    WHERE (last_updated + 43200) > ts
                ),
                batter_bb_k_results AS (
                    SELECT batter_id, so, bb
                    FROM batting_stats_standard
                    WHERE (last_updated + 43200) > ts
                )
            
            UPDATE batting_stats_advanced
            SET avg = (SELECT h / ab FROM batter_avg_results WHERE batter_id = b_id),
                obp = (SELECT (h + bb + hbp) / pa FROM batter_obp_results WHERE batter_id = b_id),
                slg = (SELECT tb / ab FROM batter_slg_results WHERE batter_id = b_id),
                ops = (SELECT slg + obp FROM batter_ops_results WHERE batter_id = b_id),
                iso = (SELECT slg - avg FROM batter_iso_results WHERE batter_id = b_id),
                babip = (SELECT hits / batted_ball_events FROM batter_babip_results WHERE batter_id = b_id),
                bb_rate = (SELECT bb / pa FROM batter_bb_rate_results WHERE batter_id = b_id),
                k_rate = (SELECT so / pa FROM batter_so_rate_results WHERE batter_id = b_id),
                bb_k = (SELECT bb / so FROM batter_bb_k_results WHERE batter_id = b_id)
            WHERE batter_id = b_id;

            DELETE FROM batting_stats_batted_ball WHERE batter_id = b_id;
            INSERT INTO batting_stats_batted_ball (batter_id, league_id, division_id, team_id, year, last_updated) VALUES (b_id, l_id, d_id, t_id, yr, ts);

            WITH 
                batter_bbe_results AS (
                    SELECT batter_id, COUNT(*) AS bbe
                    FROM trackman
                    WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ts
                    GROUP BY batter_id
                ),
                batter_gb_rate_results AS (
                    SELECT batter_id, COUNT(*), COUNT(CASE WHEN launch_angle <= 10 THEN 1 END) AS gb_events
                    FROM trackman
                    WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ts
                    GROUP BY batter_id
                ),
                batter_fb_rate_results AS (
                    SELECT batter_id, COUNT(*), COUNT(CASE WHEN launch_angle > 25 AND launch_angle <=50 THEN 1 END) AS fb_events
                    FROM trackman
                    WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ts
                    GROUP BY batter_id
                ),
                batter_ld_rate_results AS (
                    SELECT batter_id, COUNT(*), COUNT(CASE WHEN launch_angle > 10 AND launch_angle <=25 THEN 1 END) AS ld_events
                    FROM trackman
                    WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ts
                    GROUP BY batter_id
                ),
                batter_iffb_rate_results AS (
                    SELECT batter_id, COUNT(*), COUNT(CASE WHEN launch_angle > 50 THEN 1 END) AS iffb_events
                    FROM trackman
                    WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ts
                    GROUP BY batter_id
                ),
                batter_hr_fb_results AS (
                    SELECT batter_id, COUNT(CASE WHEN launch_angle > 25 AND launch_angle <=50 THEN 1 END) AS fb_events, COUNT(CASE WHEN launch_angle > 25 AND launch_angle <=50 AND result_id = 4 THEN 1 END) AS hr_events
                    FROM trackman
                    WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ts
                    GROUP BY batter_id
                ),
                batter_pull_rate_results AS (
                    SELECT trackman.batter_id, COUNT(*), COUNT(CASE WHEN (trackman.hit_bearing < -15 AND batters.batter_side_id = 1) OR (trackman.hit_bearing > 15 AND batters.batter_side_id = 2) OR
                    (trackman.hit_bearing > 15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 1) OR
                    (trackman.hit_bearing < -15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 2) THEN 1 END) AS pull_events
                    FROM trackman
                    LEFT JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
                    LEFT JOIN batters ON trackman.batter_id = batters.batter_id
                    WHERE trackman.call_id = 4 AND trackman.exit_velocity > 0 AND trackman.hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ts
                    GROUP BY trackman.batter_id
                ),
                batter_cent_rate_results AS (
                    SELECT batter_id, COUNT(*), COUNT(CASE WHEN hit_bearing >= -15 AND hit_bearing <= 15 THEN 1 END) AS cent_events
                    FROM trackman
                    WHERE call_id = 4 AND exit_velocity > 0 AND hit_bearing IS NOT NULL AND (upload_timestamp + 43200) > ts
                    GROUP BY batter_id
                ),
                batter_oppo_rate_results AS (
                    SELECT trackman.batter_id, COUNT(*), COUNT(CASE WHEN (trackman.hit_bearing > 15 AND batters.batter_side_id = 1) OR (trackman.hit_bearing < -15 AND batters.batter_side_id = 2) OR
                    (trackman.hit_bearing < -15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 1) OR
                    (trackman.hit_bearing > 15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 2) THEN 1 END) AS oppo_events
                    FROM trackman
                    LEFT JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
                    LEFT JOIN batters ON trackman.batter_id = batters.batter_id
                    WHERE trackman.call_id = 4 AND trackman.exit_velocity > 0 AND trackman.hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ts
                    GROUP BY trackman.batter_id
                ),
                batter_pull_gb_results AS (
                    SELECT trackman.batter_id, COUNT(*), COUNT(CASE WHEN (trackman.hit_bearing < -15 AND batters.batter_side_id = 1) OR (trackman.hit_bearing > 15 AND batters.batter_side_id = 2) OR
                    (trackman.hit_bearing > 15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 1) OR
                    (trackman.hit_bearing < -15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 2) THEN 1 END) AS pull_gb_events
                    FROM trackman
                    LEFT JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
                    LEFT JOIN batters ON trackman.batter_id = batters.batter_id
                    WHERE trackman.call_id = 4 AND trackman.exit_velocity > 0  AND trackman.launch_angle <= 10 AND trackman.hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ts
                    GROUP BY trackman.batter_id
                ),
                batter_cent_gb_results AS (
                    SELECT batter_id, COUNT(*), COUNT(CASE WHEN hit_bearing >= -15 AND hit_bearing <= 15 THEN 1 END) AS cent_gb_events
                    FROM trackman
                    WHERE call_id = 4 AND exit_velocity > 0 AND launch_angle <= 10 AND hit_bearing IS NOT NULL AND (upload_timestamp + 43200) > ts
                    GROUP BY batter_id
                ),
                batter_oppo_gb_results AS (
                    SELECT trackman.batter_id, COUNT(*), COUNT(CASE WHEN (trackman.hit_bearing > 15 AND batters.batter_side_id = 1) OR (trackman.hit_bearing < -15 AND batters.batter_side_id = 2) OR
                    (trackman.hit_bearing < -15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 1) OR
                    (trackman.hit_bearing > 15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 2) THEN 1 END) AS oppo_gb_events
                    FROM trackman
                    LEFT JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
                    LEFT JOIN batters ON trackman.batter_id = batters.batter_id
                    WHERE trackman.call_id = 4 AND trackman.exit_velocity > 0 AND trackman.launch_angle <= 10 AND trackman.hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ts
                    GROUP BY trackman.batter_id
                ),    
                batter_pull_fb_results AS (
                    SELECT trackman.batter_id, COUNT(*), COUNT(CASE WHEN (trackman.hit_bearing < -15 AND batters.batter_side_id = 1) OR (trackman.hit_bearing > 15 AND batters.batter_side_id = 2) OR
                    (trackman.hit_bearing > 15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 1) OR
                    (trackman.hit_bearing < -15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 2) THEN 1 END) AS pull_fb_events
                    FROM trackman
                    LEFT JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
                    LEFT JOIN batters ON trackman.batter_id = batters.batter_id
                    WHERE trackman.call_id = 4 AND trackman.exit_velocity > 0  AND trackman.launch_angle > 20 AND trackman.launch_angle <=50 AND trackman.hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ts
                    GROUP BY trackman.batter_id
                ),
                batter_cent_fb_results AS (
                    SELECT batter_id, COUNT(*), COUNT(CASE WHEN hit_bearing >= -15 AND hit_bearing <= 15 THEN 1 END) AS cent_fb_events
                    FROM trackman
                    WHERE call_id = 4 AND exit_velocity > 0 AND launch_angle > 20 AND launch_angle <=50 AND hit_bearing IS NOT NULL AND (upload_timestamp + 43200) > ts
                    GROUP BY batter_id
                ),
                batter_oppo_fb_results AS (
                    SELECT trackman.batter_id, COUNT(*), COUNT(CASE WHEN (trackman.hit_bearing > 15 AND batters.batter_side_id = 1) OR (trackman.hit_bearing < -15 AND batters.batter_side_id = 2) OR
                    (trackman.hit_bearing < -15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 1) OR
                    (trackman.hit_bearing > 15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 2) THEN 1 END) AS oppo_fb_events
                    FROM trackman
                    LEFT JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
                    LEFT JOIN batters ON trackman.batter_id = batters.batter_id
                    WHERE trackman.call_id = 4 AND trackman.exit_velocity > 0 AND trackman.launch_angle > 20 AND trackman.launch_angle <=50 AND trackman.hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ts
                    GROUP BY trackman.batter_id
                ),
                batter_soft_con_results AS (
                    SELECT batter_id, COUNT(*), COUNT(CASE WHEN exit_velocity < 70 THEN 1 END) AS soft_con_events
                    FROM trackman
                    WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ts
                    GROUP BY batter_id
                ),
                batter_med_con_results AS (
                    SELECT batter_id, COUNT(*), COUNT(CASE WHEN exit_velocity >= 70 AND exit_velocity < 95 THEN 1 END) AS med_con_events
                    FROM trackman
                    WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ts
                    GROUP BY batter_id
                ),
                batter_hard_con_results AS (
                    SELECT batter_id, COUNT(*), COUNT(CASE WHEN exit_velocity >= 95 THEN 1 END) AS hard_con_events
                    FROM trackman
                    WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ts
                    GROUP BY batter_id
                )
            
            UPDATE batting_stats_batted_ball
            SET bbe = (SELECT bbe FROM batter_bbe_results WHERE batter_id = b_id),
                gb_rate = (SELECT gb_events / bbe FROM batter_gb_rate_results WHERE batter_id = b_id),
                fb_rate = (SELECT fb_events / bbe FROM batter_fb_rate_results WHERE batter_id = b_id),
                ld_rate = (SELECT ld_events / bbe FROM batter_ld_rate_results WHERE batter_id = b_id),
                iffb_rate = (SELECT iffb_events / bbe FROM batter_iffb_rate_results WHERE batter_id = b_id),
                hr_fb = (SELECT hr_events / fb_events FROM batter_hr_fb_results WHERE batter_id = b_id),
                pull_rate = (SELECT pull_events / bbe FROM batter_pull_rate_results WHERE batter_id = b_id),
                cent_rate = (SELECT cent_events / bbe FROM batter_cent_rate_results WHERE batter_id = b_id),
                oppo_rate = (SELECT oppo_events / bbe FROM batter_oppo_rate_results WHERE batter_id = b_id),
                pull_gb_rate = (SELECT pull_gb_events / gb_events FROM batter_pull_gb_results WHERE batter_id = b_id),
                cent_gb_rate = (SELECT cent_gb_events / gb_events FROM batter_cent_gb_results WHERE batter_id = b_id),
                oppo_gb_rate = (SELECT oppo_gb_events / gb_events FROM batter_oppo_gb_results WHERE batter_id = b_id),
                pull_fb_rate = (SELECT pull_fb_events / fb_events FROM batter_pull_fb_results WHERE batter_id = b_id),
                cent_fb_rate = (SELECT cent_fb_events / fb_events FROM batter_cent_fb_results WHERE batter_id = b_id),
                oppo_fb_rate = (SELECT oppo_fb_events / fb_events FROM batter_oppo_fb_results WHERE batter_id = b_id),
                soft_con_rate = (SELECT soft_con_events / bbe FROM batter_soft_con_results WHERE batter_id = b_id),
                med_con_rate = (SELECT med_con_events / bbe FROM batter_med_con_results WHERE batter_id = b_id),
                hard_con_rate = (SELECT hard_con_events / bbe FROM batter_hard_con_results WHERE batter_id = b_id)
            WHERE batter_id = b_id;
                          
            DELETE FROM batting_stats_discipline WHERE batter_id = b_id;
            INSERT INTO batting_stats_batted_ball (batter_id, league_id, division_id, team_id, year, last_updated) VALUES (b_id, l_id, d_id, t_id, yr, ts);
            
            WITH
                batter_o_swing_results AS (
                    SELECT batter_id, COUNT(*) AS ooz_events, COUNT(CASE WHEN call_id = 2 OR call_id = 3 OR call_id = 4 THEN 1 END) AS swing_events
                    FROM trackman
                    WHERE (location_side < -0.7508 OR location_side > 0.7508 OR location_height < 1.5942 OR location_height > 3.6033) AND (upload_timestamp + 43200) > ts
                    GROUP BY batter_id
                ),
                batter_z_swing_results AS (
                    SELECT batter_id, COUNT(*) AS izz_events, COUNT(CASE WHEN call_id = 2 OR call_id = 3 OR call_id = 4 THEN 1 END) AS swing_events
                    FROM trackman
                    WHERE (location_side >= -0.7508 AND location_side <= 0.7508 AND location_height >= 1.5942 AND location_height <= 3.6033) AND (upload_timestamp + 43200) > ts
                    GROUP BY batter_id
                ),
                batter_swing_rate_results AS (
                    SELECT batter_id, COUNT(*) AS total_events, COUNT(CASE WHEN call_id = 2 OR call_id = 3 OR call_id = 4 THEN 1 END) AS swing_events
                    FROM trackman
                    WHERE (upload_timestamp + 43200) > ts
                    GROUP BY batter_id
                ),
                batter_o_con_results AS (
                    SELECT batter_id, COUNT(*) AS ooz_events, COUNT(CASE WHEN call_id = 2 OR call_id = 3 OR call_id = 4 THEN 1 END) AS swing_events, COUNT(CASE WHEN call_id = 3 OR call_id = 4 THEN 1 END) AS con_events
                    FROM trackman
                    WHERE (location_side < -0.7508 OR location_side > 0.7508 OR location_height < 1.5942 OR location_height > 3.6033) AND (upload_timestamp + 43200) > ts
                    GROUP BY batter_id
                ),
                batter_z_con_results AS (
                    SELECT batter_id, COUNT(*) AS izz_events, COUNT(CASE WHEN call_id = 2 OR call_id = 3 OR call_id = 4 THEN 1 END) AS swing_events, COUNT(CASE WHEN call_id = 3 OR call_id = 4 THEN 1 END) AS con_events
                    FROM trackman
                    WHERE (location_side >= -0.7508 AND location_side <= 0.7508 AND location_height >= 1.5942 AND location_height <= 3.6033) AND (upload_timestamp + 43200) > ts
                    GROUP BY batter_id
                ),
                batter_con_rate_results AS (
                    SELECT batter_id, COUNT(*) AS total_events, COUNT(CASE WHEN call_id = 2 OR call_id = 3 OR call_id = 4 THEN 1 END) AS swing_events, COUNT(CASE WHEN call_id = 3 OR call_id = 4 THEN 1 END) AS con_events
                    FROM trackman
                    WHERE (upload_timestamp + 43200) > ts
                    GROUP BY batter_id
                ),
                batter_izz_rate_results AS (
                    SELECT batter_id, COUNT(*) AS total_events, COUNT(CASE WHEN location_side >= -0.7508 AND location_side <= 0.7508 AND location_height >= 1.5942 AND location_height <= 3.6033 THEN 1 END) AS izz_events
                    FROM trackman
                    WHERE (upload_timestamp + 43200) > ts
                    GROUP BY batter_id
                )
            
            UPDATE batting_stats_discipline
            SET o_swing = (SELECT swing_events / ooz_events FROM batter_o_swing_results WHERE batter_id = b_id),
            z_swing = (SELECT swing_events / izz_events FROM batter_z_swing_results WHERE batter_id = b_id),
            swing_rate = (SELECT swing_events / total_events FROM batter_swing_rate_results WHERE batter_id = b_id),
            o_contact = (SELECT con_events / swing_events FROM batter_o_con_results WHERE batter_id = b_id),
            z_contact = (SELECT con_events / swing_events FROM batter_z_con_results WHERE batter_id = b_id),
            contact_rate = (SELECT con_events / swing_events FROM batter_con_rate_results WHERE batter_id = b_id),
            zone_rate = (SELECT izz_events / total_events FROM batter_izz_rate_results WHERE batter_id = b_id)

        END;''')
        
    def removeGame(self):

        conn = sqlite3.connect('rylar_baseball.db')
        cur = conn.cursor()

        cur.execute("SELECT game_id FROM games WHERE trackman_id = ?", (self.trackman_id,))
        game_id = cur.fetchone()[0]

        cur.execute("DELETE FROM games WHERE game_id = ?", (game_id,))
        cur.execute("DELETE FROM trackman WHERE game_id = ?", (game_id,))
        
        print('Game deleted from database!')
        conn.commit()
        cur.close()

    def updateStats(self):

        print('...updating stats...')

        conn = sqlite3.connect('rylar_baseball.db')
        cur = conn.cursor()

        #NEXT MOVE IS TO PULL OUT THE DATA WE ARE GOING TO USE ONCE. SNAG EVERY PLAYER, PITCH, ETC AND WRITE TO A TEMPORARY TABLE
        #DONT NEED TO LOCATE EVERY PITCH OVER AND OVER AND OVER AGAIN.

        #Hitters 

        # Standard
        
        cur.execute("BEGIN")

        #Get player_ids for players involved in a game uploaded within the last 12 hours
        cur.execute('''SELECT trackman.batter_id, trackman.league_id, trackman.division_id, batters.team_id, teams.year trackman
                    FROM trackman
                    LEFT JOIN batters ON trackman.batter_id = batters.batter_id
                    LEFT JOIN teams ON batters.team_id = teams.team_id
                    WHERE (trackman.upload_timestamp + 43200) > ?
                    GROUP BY trackman.batter_id, teams.year''', (self.timestamp,))
        tupIDs = [tup + (self.timestamp,) for tup in cur.fetchall()]
        batterIDs = [tup[0] for tup in tupIDs]

        #Clear standard table of recent players if they are already listed
        cur.executemany('DELETE FROM batting_stats_standard WHERE batter_id = ?', [(id,) for id in batterIDs])
        
        #Insert every player id into the table that played in recently uploaded games
        cur.executemany('INSERT INTO batting_stats_standard (batter_id, league_id, division_id, team_id, year, last_updated) VALUES (?,?,?,?,?,?)', (tupIDs))
        
        #Get games by batter id
        cur.execute('SELECT batter_id, COUNT(DISTINCT game_id) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY batter_id', (self.timestamp,))
        gByID = dict(cur.fetchall())
        cur.executemany('UPDATE batting_stats_standard SET g = ? WHERE batter_id = ?', ([(gByID.get(id, 0), id) for id in batterIDs]))

        #Get at bats by batter id (chat gpt helped with the partition and row_number, but it looks like it works with a couple fixes!)
        cur.execute('''SELECT batter_id, COUNT(*)
                    FROM (SELECT *, ROW_NUMBER() OVER (PARTITION BY batter_id, game_id, inning, pa_of_inning ORDER BY pitch_num DESC) AS row_num FROM trackman)
                    WHERE (row_num = 1) AND (k_or_bb_id != ?) AND (result_id != ?) AND (call_id != ?) AND (upload_timestamp + 43200) > ?
                    GROUP BY batter_id''', ('3', '7', '6', self.timestamp))
        abByID = dict(cur.fetchall())
        cur.executemany('UPDATE batting_stats_standard SET ab = ? WHERE batter_id = ?', ([(abByID.get(id, 0), id) for id in batterIDs]))

        #Get plate appearances by batter id
        cur.execute('SELECT batter_id, COUNT(DISTINCT (game_id || inning || pa_of_inning)) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY batter_id', (self.timestamp,))
        paByID = dict(cur.fetchall())
        cur.executemany('UPDATE batting_stats_standard SET pa = ? WHERE batter_id = ?', ([(paByID.get(id, 0), id) for id in batterIDs]))

        #Get hits by batter id
        cur.execute('SELECT batter_id, COUNT(*) FROM trackman WHERE result_id <= 4 AND (upload_timestamp + 43200) > ? GROUP BY batter_id', (self.timestamp,))
        hByID = dict(cur.fetchall())
        cur.executemany('UPDATE batting_stats_standard SET h = ? WHERE batter_id = ?', ([(hByID.get(id, 0), id) for id in batterIDs]))

        #Get doubles by batter id
        cur.execute('SELECT batter_id, COUNT(*) FROM trackman WHERE result_id = 2 AND (upload_timestamp + 43200) > ? GROUP BY batter_id', (self.timestamp,))
        _2bByID = dict(cur.fetchall())
        cur.executemany('UPDATE batting_stats_standard SET _2b = ? WHERE batter_id = ?', ([(_2bByID.get(id, 0), id) for id in batterIDs]))

        #Get triples by batter id
        cur.execute('SELECT batter_id, COUNT(*) FROM trackman WHERE result_id = 3 AND (upload_timestamp + 43200) > ? GROUP BY batter_id', (self.timestamp,))
        _3bByID = dict(cur.fetchall())
        cur.executemany('UPDATE batting_stats_standard SET _3b = ? WHERE batter_id = ?', ([(_3bByID.get(id, 0), id) for id in batterIDs]))

        #Get homers by batter id
        cur.execute('SELECT batter_id, COUNT(*) FROM trackman WHERE result_id = 4 AND (upload_timestamp + 43200) > ? GROUP BY batter_id', (self.timestamp,))
        hrByID = dict(cur.fetchall())
        cur.executemany('UPDATE batting_stats_standard SET hr = ? WHERE batter_id = ?', ([(hrByID.get(id, 0), id) for id in batterIDs]))

        #Get rbis by batter id
        cur.execute('SELECT batter_id, SUM(runs_scored) FROM trackman WHERE runs_scored >= 1 AND outs_made < 2 AND (upload_timestamp + 43200) > ? GROUP BY batter_id', (self.timestamp,))
        rbiByID = dict(cur.fetchall())
        cur.executemany('UPDATE batting_stats_standard SET rbi = ? WHERE batter_id = ?', ([(rbiByID.get(id, 0), id) for id in batterIDs]))

        #Get walks by batter id
        cur.execute('SELECT batter_id, COUNT(*) FROM trackman WHERE k_or_bb_id = 3 AND call_id != 9 AND (upload_timestamp + 43200) > ? GROUP BY batter_id', (self.timestamp,))
        bbByID = dict(cur.fetchall())
        cur.executemany('UPDATE batting_stats_standard SET bb = ? WHERE batter_id = ?', ([(bbByID.get(id, 0), id) for id in batterIDs]))

        #Get intentional walks by batter id
        cur.execute('SELECT batter_id, COUNT(*) FROM trackman WHERE k_or_bb_id = 3 AND call_id = 9 AND (upload_timestamp + 43200) > ? GROUP BY batter_id', (self.timestamp,))
        ibbByID = dict(cur.fetchall())
        cur.executemany('UPDATE batting_stats_standard SET ibb = ? WHERE batter_id = ?', ([(ibbByID.get(id, 0), id) for id in batterIDs]))

        #Get strikeouts by batter id
        cur.execute('SELECT batter_id, COUNT(*) FROM trackman WHERE k_or_bb_id = 2 AND (upload_timestamp + 43200) > ? GROUP BY batter_id', (self.timestamp,))
        soByID = dict(cur.fetchall())
        cur.executemany('UPDATE batting_stats_standard SET so = ? WHERE batter_id = ?', ([(soByID.get(id, 0), id) for id in batterIDs]))

        #Get hit by pitch by batter id
        cur.execute('SELECT batter_id, COUNT(*) FROM trackman WHERE call_id = 6 AND (upload_timestamp + 43200) > ? GROUP BY batter_id', (self.timestamp,))
        hbpByID = dict(cur.fetchall())
        cur.executemany('UPDATE batting_stats_standard SET hbp = ? WHERE batter_id = ?', ([(hbpByID.get(id, 0), id) for id in batterIDs]))

        #Get sacrifices by batter id
        cur.execute('SELECT batter_id, COUNT(*) FROM trackman WHERE result_id = 7 AND (upload_timestamp + 43200) > ? GROUP BY batter_id', (self.timestamp,))
        sacByID = dict(cur.fetchall())
        cur.executemany('UPDATE batting_stats_standard SET sac = ? WHERE batter_id = ?', ([(sacByID.get(id, 0), id) for id in batterIDs]))

        #Get grounded into double plays by batter id
        cur.execute('SELECT batter_id, COUNT(*) FROM trackman WHERE outs_made >= 2 AND hit_type_id = 4 AND (upload_timestamp + 43200) > ? GROUP BY batter_id', (self.timestamp,))
        gdpByID = dict(cur.fetchall())
        cur.executemany('UPDATE batting_stats_standard SET gdp = ? WHERE batter_id = ?', ([(gdpByID.get(id, 0), id) for id in batterIDs]))

        #Get total bases by batter id
        cur.execute('SELECT batter_id, h, _2b, _3b, hr FROM batting_stats_standard WHERE (last_updated + 43200) > ?', (self.timestamp,))
        tupsByID = cur.fetchall()
        tbByID = {tup[0] : ((tup[1]-tup[2]-tup[3]-tup[4]) + 2*tup[2] + 3*tup[3] + 4*tup[4]) for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_standard SET tb = ? WHERE batter_id = ?', ([(tbByID.get(id, 0), id) for id in batterIDs]))

        # Statcast xBA, xSLG, xwOBA

        #Clear statcast table of recent players if they are already listed
        cur.executemany('DELETE FROM batting_stats_statcast WHERE batter_id = ?', [(id,) for id in batterIDs])
        
        #Insert every player id into the table that played in recently uploaded games
        cur.executemany('INSERT INTO batting_stats_statcast (batter_id, league_id, division_id, team_id, year, last_updated) VALUES (?,?,?,?,?,?)', (tupIDs))

        #Get every batted ball event by batter id
        cur.execute('SELECT batter_id, COUNT(*) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY batter_id', (self.timestamp,))
        bbeByID = dict(cur.fetchall())
        cur.executemany('UPDATE batting_stats_statcast SET bbe = ? WHERE batter_id = ?', ([(bbeByID.get(id, 0), id) for id in batterIDs]))

        #Get average exit velo by batter id
        cur.execute('SELECT batter_id, AVG(exit_velocity) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY batter_id', (self.timestamp,))
        avg_evByID = dict(cur.fetchall())
        cur.executemany('UPDATE batting_stats_statcast SET avg_ev = ? WHERE batter_id = ?', ([(avg_evByID.get(id, None), id) for id in batterIDs]))

        #Get max exit velo by batter id
        cur.execute('SELECT batter_id, MAX(exit_velocity) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY batter_id', (self.timestamp,))
        max_evByID = dict(cur.fetchall())
        cur.executemany('UPDATE batting_stats_statcast SET max_ev = ? WHERE batter_id = ?', ([(max_evByID.get(id, None), id) for id in batterIDs]))

        #Get average launch angle by batter id
        cur.execute('SELECT batter_id, AVG(launch_angle) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY batter_id', (self.timestamp,))
        avg_laByID = dict(cur.fetchall())
        cur.executemany('UPDATE batting_stats_statcast SET avg_la = ? WHERE batter_id = ?', ([(avg_laByID.get(id, None), id) for id in batterIDs]))

        #Get barrels by batter id (probably need a league specific definition of barrel) 
        cur.execute('SELECT batter_id, exit_velocity, launch_angle FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? ', (self.timestamp,))
        tupsByID = cur.fetchall()

        brlsByID = Counter()
        for tup in tupsByID:
            id = tup[0]
            exit_velocity = float(tup[1])
            launch_angle = float(tup[2])
            barrel = False
            if exit_velocity >= 98.0:
                if exit_velocity <= 99.0:
                    if launch_angle >= 26.0 and launch_angle <= 30.0:
                        barrel = True
                elif exit_velocity <= 100.0:
                    if launch_angle >= 25.0 and launch_angle <= 31.0:
                        barrel = True
                #Not a perfect representation of a barrel, but pretty close
                else:
                    range_growth = (exit_velocity - 100.0) * 1.2
                    high_angle = min(31.0 + range_growth, 50.0)
                    low_angle = max(25.0 - range_growth, 8.0)
                    if launch_angle >= low_angle and launch_angle <= high_angle:
                        barrel = True
            if barrel:
                brlsByID[id] += 1
        brlsByID = dict(brlsByID)
        cur.executemany('UPDATE batting_stats_statcast SET brls = ? WHERE batter_id = ?', ([(brlsByID.get(id, 0), id) for id in batterIDs]))
        
        #Get barrel rate by batter id
        cur.execute('SELECT batter_id, bbe, brls FROM batting_stats_statcast WHERE (last_updated + 43200) > ? ', (self.timestamp,))
        tupsByID = cur.fetchall()
        brl_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_statcast SET brl_rate = ? WHERE batter_id = ?', ([(brl_rateByID.get(id, None), id) for id in batterIDs]))
        
        #Get hard hits by batter id
        cur.execute('SELECT batter_id, COUNT(*) FROM trackman WHERE call_id = 4 AND exit_velocity >= 95 AND (upload_timestamp + 43200) > ? GROUP BY batter_id', (self.timestamp,))
        hhByID = dict(cur.fetchall())
        cur.executemany('UPDATE batting_stats_statcast SET hh = ? WHERE batter_id = ?', ([(hhByID.get(id, 0), id) for id in batterIDs]))
        
        #Get hard hit rate by batter id
        cur.execute('SELECT batter_id, bbe, hh FROM batting_stats_statcast WHERE (last_updated + 43200) > ?', (self.timestamp,))
        tupsByID = cur.fetchall()
        hh_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_statcast SET hh_rate = ? WHERE batter_id = ?', ([(hh_rateByID.get(id, None), id) for id in batterIDs]))
        
        # Advanced wRC, wRAA, wOBA

        #Clear advanced table of recent players if they are already listed
        cur.executemany('DELETE FROM batting_stats_advanced WHERE batter_id = ?', [(id,) for id in batterIDs])
        
        #Insert every player id into the table that played in recently uploaded games
        cur.executemany('INSERT INTO batting_stats_advanced (batter_id, league_id, division_id, team_id, year, last_updated) VALUES (?,?,?,?,?,?)', (tupIDs))

        #Get average by batter id
        cur.execute('SELECT batter_id, ab, h FROM batting_stats_standard WHERE (last_updated + 43200) > ?', (self.timestamp,))
        tupsByID = cur.fetchall()
        avgByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_advanced SET avg = ? WHERE batter_id = ?', ([(avgByID.get(id, 0), id) for id in batterIDs]))

        #Get obp by batter id
        cur.execute('SELECT batter_id, pa, h, bb, hbp FROM batting_stats_standard WHERE (last_updated + 43200) > ?', (self.timestamp,))
        tupsByID = cur.fetchall()
        obpByID = {tup[0] : (tup[2] + tup[3] + tup[4]) / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_advanced SET obp = ? WHERE batter_id = ?', ([(obpByID.get(id, 0), id) for id in batterIDs]))

        #Get slg by batter id
        cur.execute('SELECT batter_id, ab, tb FROM batting_stats_standard WHERE (last_updated + 43200) > ?', (self.timestamp,))
        tupsByID = cur.fetchall()
        slgByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_advanced SET slg = ? WHERE batter_id = ?', ([(slgByID.get(id, 0), id) for id in batterIDs]))
        
        #Get ops by batter id
        cur.execute('SELECT batter_id, slg, obp FROM batting_stats_advanced WHERE (last_updated + 43200) > ?', (self.timestamp,))
        tupsByID = cur.fetchall()
        opsByID = {tup[0] : tup[2] + tup[1] for tup in tupsByID}
        cur.executemany('UPDATE batting_stats_advanced SET ops = ? WHERE batter_id = ?', ([(opsByID.get(id, 0), id) for id in batterIDs]))
        
        #Get iso by batter id
        cur.execute('SELECT batter_id, avg, slg FROM batting_stats_advanced WHERE (last_updated + 43200) > ?', (self.timestamp,))
        tupsByID = cur.fetchall()
        isoByID = {tup[0] : tup[2] - tup[1] for tup in tupsByID}
        cur.executemany('UPDATE batting_stats_advanced SET iso = ? WHERE batter_id = ?', ([(isoByID.get(id, 0), id) for id in batterIDs]))
        
        #Get babip by batter id
        cur.execute('SELECT batter_id, COUNT(*), COUNT(CASE WHEN result_id <= 4 THEN 1 END) FROM trackman WHERE call_id = 4 AND (upload_timestamp + 43200) > ? GROUP BY batter_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        babipByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_advanced SET babip = ? WHERE batter_id = ?', ([(babipByID.get(id, None), id) for id in batterIDs]))

        #Get walk rate by batter id
        cur.execute('SELECT batter_id, pa, bb FROM batting_stats_standard WHERE (last_updated + 43200) > ?', (self.timestamp,))
        tupsByID = cur.fetchall()
        bb_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_advanced SET bb_rate = ? WHERE batter_id = ?', ([(bb_rateByID.get(id, None), id) for id in batterIDs]))

        #Get strikeout rate by batter id
        cur.execute('SELECT batter_id, pa, so FROM batting_stats_standard WHERE (last_updated + 43200) > ?', (self.timestamp,))
        tupsByID = cur.fetchall()
        so_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_advanced SET k_rate = ? WHERE batter_id = ?', ([(so_rateByID.get(id, None), id) for id in batterIDs]))

        #Get bb/k by batter id
        cur.execute('SELECT batter_id, so, bb FROM batting_stats_standard WHERE (last_updated + 43200) > ?', (self.timestamp,))
        tupsByID = cur.fetchall()
        bb_kByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_advanced SET bb_k = ? WHERE batter_id = ?', ([(bb_kByID.get(id, None), id) for id in batterIDs]))
        
        # Batted Ball

        #Clear batted ball table of recent players if they are already listed
        cur.executemany('DELETE FROM batting_stats_batted_ball WHERE batter_id = ?', [(id,) for id in batterIDs])
        
        #Insert every player id into the table that played in recently uploaded games
        cur.executemany('INSERT INTO batting_stats_batted_ball (batter_id, league_id, division_id, team_id, year, last_updated) VALUES (?,?,?,?,?,?)', (tupIDs))
        
        #Insert every batted ball event by batter id
        cur.executemany('UPDATE batting_stats_batted_ball SET bbe = ? WHERE batter_id = ?', ([(bbeByID.get(id, 0), id) for id in batterIDs]))

        #Get every ground ball rate by batter id
        cur.execute('SELECT batter_id, COUNT(*), COUNT(CASE WHEN launch_angle <= 10 THEN 1 END) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY batter_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        gb_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_batted_ball SET gb_rate = ? WHERE batter_id = ?', ([(gb_rateByID.get(id, None), id) for id in batterIDs]))

        #Get every fly ball rate by batter id
        cur.execute('SELECT batter_id, COUNT(*), COUNT(CASE WHEN launch_angle > 25 AND launch_angle <=50 THEN 1 END) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY batter_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        fb_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_batted_ball SET fb_rate = ? WHERE batter_id = ?', ([(fb_rateByID.get(id, None), id) for id in batterIDs]))

        #Get every line drive rate by batter id
        cur.execute('SELECT batter_id, COUNT(*), COUNT(CASE WHEN launch_angle > 10 AND launch_angle <=25 THEN 1 END) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY batter_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        ld_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_batted_ball SET ld_rate = ? WHERE batter_id = ?', ([(ld_rateByID.get(id, None), id) for id in batterIDs]))

        #Get every infield fly ball rate rate by batter id
        cur.execute('SELECT batter_id, COUNT(*), COUNT(CASE WHEN launch_angle > 50 THEN 1 END) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY batter_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        iffb_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_batted_ball SET iffb_rate = ? WHERE batter_id = ?', ([(iffb_rateByID.get(id, None), id) for id in batterIDs]))

        #Get every home run per fly ball rate by batter id
        cur.execute('SELECT batter_id, COUNT(CASE WHEN launch_angle > 25 AND launch_angle <=50 THEN 1 END), COUNT(CASE WHEN launch_angle > 25 AND launch_angle <=50 AND result_id = 4 THEN 1 END) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY batter_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        hr_fbByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_batted_ball SET hr_fb = ? WHERE batter_id = ?', ([(hr_fbByID.get(id, None), id) for id in batterIDs]))

        #Get every pull rate by batter id
        cur.execute('''SELECT trackman.batter_id, COUNT(*),
        COUNT(CASE WHEN (trackman.hit_bearing < -15 AND batters.batter_side_id = 1) OR (trackman.hit_bearing > 15 AND batters.batter_side_id = 2) OR
        (trackman.hit_bearing > 15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 1) OR
        (trackman.hit_bearing < -15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 2) THEN 1 END)
        FROM trackman
        LEFT JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
        LEFT JOIN batters ON trackman.batter_id = batters.batter_id
        WHERE trackman.call_id = 4 AND trackman.exit_velocity > 0 AND trackman.hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ? GROUP BY trackman.batter_id''', (self.timestamp,))
        tupsByID = cur.fetchall()
        pull_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_batted_ball SET pull_rate = ? WHERE batter_id = ?', ([(pull_rateByID.get(id, None), id) for id in batterIDs]))

        #Get every center rate by batter id
        cur.execute('''SELECT batter_id, COUNT(*),
        COUNT(CASE WHEN hit_bearing >= -15 AND hit_bearing <= 15 THEN 1 END)
        FROM trackman
        WHERE call_id = 4 AND exit_velocity > 0 AND hit_bearing IS NOT NULL AND (upload_timestamp + 43200) > ? GROUP BY batter_id''', (self.timestamp,))
        tupsByID = cur.fetchall()
        cent_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_batted_ball SET cent_rate = ? WHERE batter_id = ?', ([(cent_rateByID.get(id, None), id) for id in batterIDs]))

        #Get every oppo rate by batter id
        cur.execute('''SELECT trackman.batter_id, COUNT(*),
        COUNT(CASE WHEN (trackman.hit_bearing > 15 AND batters.batter_side_id = 1) OR (trackman.hit_bearing < -15 AND batters.batter_side_id = 2) OR
        (trackman.hit_bearing < -15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 1) OR
        (trackman.hit_bearing > 15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 2) THEN 1 END)
        FROM trackman
        LEFT JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
        LEFT JOIN batters ON trackman.batter_id = batters.batter_id
        WHERE trackman.call_id = 4 AND trackman.exit_velocity > 0 AND trackman.hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ? GROUP BY trackman.batter_id''', (self.timestamp,))
        tupsByID = cur.fetchall()
        oppo_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_batted_ball SET oppo_rate = ? WHERE batter_id = ?', ([(oppo_rateByID.get(id, None), id) for id in batterIDs]))

        #Get every pull per gb rate by batter id
        cur.execute('''SELECT trackman.batter_id, COUNT(*),
        COUNT(CASE WHEN (trackman.hit_bearing < -15 AND batters.batter_side_id = 1) OR (trackman.hit_bearing > 15 AND batters.batter_side_id = 2) OR
        (trackman.hit_bearing > 15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 1) OR
        (trackman.hit_bearing < -15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 2) THEN 1 END)
        FROM trackman
        LEFT JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
        LEFT JOIN batters ON trackman.batter_id = batters.batter_id
        WHERE trackman.call_id = 4 AND trackman.exit_velocity > 0  AND trackman.launch_angle <= 10 AND trackman.hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ? GROUP BY trackman.batter_id''', (self.timestamp,))
        tupsByID = cur.fetchall()
        pull_gbByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_batted_ball SET pull_gb_rate = ? WHERE batter_id = ?', ([(pull_gbByID.get(id, None), id) for id in batterIDs]))

        #Get every center per gb rate by batter id
        cur.execute('''SELECT batter_id, COUNT(*),
        COUNT(CASE WHEN hit_bearing >= -15 AND hit_bearing <= 15 THEN 1 END)
        FROM trackman
        WHERE call_id = 4 AND exit_velocity > 0 AND launch_angle <= 10 AND hit_bearing IS NOT NULL AND (upload_timestamp + 43200) > ? GROUP BY batter_id''', (self.timestamp,))
        tupsByID = cur.fetchall()
        cent_gbByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_batted_ball SET cent_gb_rate = ? WHERE batter_id = ?', ([(cent_gbByID.get(id, None), id) for id in batterIDs]))

        #Get every oppo per gb rate by batter id
        cur.execute('''SELECT trackman.batter_id, COUNT(*),
        COUNT(CASE WHEN (trackman.hit_bearing > 15 AND batters.batter_side_id = 1) OR (trackman.hit_bearing < -15 AND batters.batter_side_id = 2) OR
        (trackman.hit_bearing < -15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 1) OR
        (trackman.hit_bearing > 15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 2) THEN 1 END)
        FROM trackman
        LEFT JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
        LEFT JOIN batters ON trackman.batter_id = batters.batter_id
        WHERE trackman.call_id = 4 AND trackman.exit_velocity > 0 AND trackman.launch_angle <= 10 AND trackman.hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ? GROUP BY trackman.batter_id''', (self.timestamp,))
        tupsByID = cur.fetchall()
        oppo_gbByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_batted_ball SET oppo_gb_rate = ? WHERE batter_id = ?', ([(oppo_gbByID.get(id, None), id) for id in batterIDs]))

        #Get every pull per fb rate by batter id (use lower, lower bound for launch angle for purpose of stat--to be used for outfield positioning)
        cur.execute('''SELECT trackman.batter_id, COUNT(*),
        COUNT(CASE WHEN (trackman.hit_bearing < -15 AND batters.batter_side_id = 1) OR (trackman.hit_bearing > 15 AND batters.batter_side_id = 2) OR
        (trackman.hit_bearing > 15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 1) OR
        (trackman.hit_bearing < -15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 2) THEN 1 END)
        FROM trackman
        LEFT JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
        LEFT JOIN batters ON trackman.batter_id = batters.batter_id
        WHERE trackman.call_id = 4 AND trackman.exit_velocity > 0  AND trackman.launch_angle > 20 AND trackman.launch_angle <=50 AND trackman.hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ? GROUP BY trackman.batter_id''', (self.timestamp,))
        tupsByID = cur.fetchall()
        pull_fbByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_batted_ball SET pull_fb_rate = ? WHERE batter_id = ?', ([(pull_fbByID.get(id, None), id) for id in batterIDs]))

        #Get every center per fb rate by batter id (use lower, lower bound for launch angle for purpose of stat--to be used for outfield positioning)
        cur.execute('''SELECT batter_id, COUNT(*),
        COUNT(CASE WHEN hit_bearing >= -15 AND hit_bearing <= 15 THEN 1 END)
        FROM trackman
        WHERE call_id = 4 AND exit_velocity > 0 AND launch_angle > 20 AND launch_angle <=50 AND hit_bearing IS NOT NULL AND (upload_timestamp + 43200) > ? GROUP BY batter_id''', (self.timestamp,))
        tupsByID = cur.fetchall()
        cent_fbByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_batted_ball SET cent_fb_rate = ? WHERE batter_id = ?', ([(cent_fbByID.get(id, None), id) for id in batterIDs]))

        #Get every oppo per fb rate by batter id (use lower, lower bound for launch angle for purpose of stat--to be used for outfield positioning)
        cur.execute('''SELECT trackman.batter_id, COUNT(*),
        COUNT(CASE WHEN (trackman.hit_bearing > 15 AND batters.batter_side_id = 1) OR (trackman.hit_bearing < -15 AND batters.batter_side_id = 2) OR
        (trackman.hit_bearing < -15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 1) OR
        (trackman.hit_bearing > 15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 2) THEN 1 END)
        FROM trackman
        LEFT JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
        LEFT JOIN batters ON trackman.batter_id = batters.batter_id
        WHERE trackman.call_id = 4 AND trackman.exit_velocity > 0 AND trackman.launch_angle > 20 AND trackman.launch_angle <=50 AND trackman.hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ? GROUP BY trackman.batter_id''', (self.timestamp,))
        tupsByID = cur.fetchall()
        oppo_fbByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_batted_ball SET oppo_fb_rate = ? WHERE batter_id = ?', ([(oppo_fbByID.get(id, None), id) for id in batterIDs]))

        #Get soft con rate by batter id
        cur.execute('SELECT batter_id, COUNT(*), COUNT(CASE WHEN exit_velocity < 70 THEN 1 END) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY batter_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        soft_conbyID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_batted_ball SET soft_con_rate = ? WHERE batter_id = ?', ([(soft_conbyID.get(id, None), id) for id in batterIDs]))

        #Get med con rate by batter id
        cur.execute('SELECT batter_id, COUNT(*), COUNT(CASE WHEN exit_velocity >= 70 AND exit_velocity < 95 THEN 1 END) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY batter_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        med_conbyID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_batted_ball SET med_con_rate = ? WHERE batter_id = ?', ([(med_conbyID.get(id, None), id) for id in batterIDs]))

        #Get hard con rate by batter id
        cur.execute('SELECT batter_id, COUNT(*), COUNT(CASE WHEN exit_velocity >= 95 THEN 1 END) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY batter_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        hard_conbyID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_batted_ball SET hard_con_rate = ? WHERE batter_id = ?', ([(hard_conbyID.get(id, None), id) for id in batterIDs]))
        
        # Discipline

        #Clear discipline table of recent players if they are already listed
        cur.executemany('DELETE FROM batting_stats_batted_ball WHERE batter_id = ?', [(id,) for id in batterIDs])
        
        #Insert every player id into the table that played in recently uploaded games
        cur.executemany('INSERT INTO batting_stats_batted_ball (batter_id, league_id, division_id, team_id, year, last_updated) VALUES (?,?,?,?,?,?)', (tupIDs))
        
        #Get outside of zone swing rate by batter id
        #Get balls based on universal strike zone
        cur.execute('SELECT batter_id, COUNT(*), COUNT(CASE WHEN call_id = 2 OR call_id = 3 OR call_id = 4 THEN 1 END) FROM trackman WHERE (upload_timestamp + 43200) > ? AND (location_side < -0.7508 OR location_side > 0.7508 OR location_height < 1.5942 OR location_height > 3.6033) GROUP BY batter_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        o_swingbyID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_discipline SET o_swing = ? WHERE batter_id = ?', ([(o_swingbyID.get(id, None), id) for id in batterIDs]))

        #Get inside of zone swing rate by batter id
        #Get strikes based on universal strike zone
        cur.execute('SELECT batter_id, COUNT(*), COUNT(CASE WHEN call_id = 2 OR call_id = 3 OR call_id = 4 THEN 1 END) FROM trackman WHERE (upload_timestamp + 43200) > ? AND location_side >= -0.7508 AND location_side <= 0.7508 AND location_height >= 1.5942 AND location_height <= 3.6033 GROUP BY batter_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        z_swingbyID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_discipline SET z_swing = ? WHERE batter_id = ?', ([(z_swingbyID.get(id, None), id) for id in batterIDs]))

        #Get swing rate by batter id
        cur.execute('SELECT batter_id, COUNT(*), COUNT(CASE WHEN call_id = 2 OR call_id = 3 OR call_id = 4 THEN 1 END) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY batter_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        swing_ratebyID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_discipline SET swing_rate = ? WHERE batter_id = ?', ([(swing_ratebyID.get(id, None), id) for id in batterIDs]))
  
        #Get outside of zone contact rate by batter id
        #Get balls based on universal strike zone
        cur.execute('SELECT batter_id, COUNT(CASE WHEN call_id = 2 OR call_id = 3 OR call_id = 4 THEN 1 END), COUNT(CASE WHEN call_id = 3 OR call_id = 4 THEN 1 END) FROM trackman WHERE (upload_timestamp + 43200) > ? AND location_side < -0.7508 OR location_side > 0.7508 OR location_height < 1.5942 OR location_height > 3.6033 GROUP BY batter_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        o_contactbyID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_discipline SET o_contact = ? WHERE batter_id = ?', ([(o_contactbyID.get(id, None), id) for id in batterIDs]))

        #Get inside of zone contact rate by batter id
        #Get strikes based on universal strike zone
        cur.execute('SELECT batter_id, COUNT(CASE WHEN call_id = 2 OR call_id = 3 OR call_id = 4 THEN 1 END), COUNT(CASE WHEN call_id = 3 OR call_id = 4 THEN 1 END) FROM trackman WHERE (upload_timestamp + 43200) > ? AND location_side >= -0.7508 AND location_side <= 0.7508 AND location_height >= 1.5942 AND location_height <= 3.6033 GROUP BY batter_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        z_contactbyID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_discipline SET z_contact = ? WHERE batter_id = ?', ([(z_contactbyID.get(id, None), id) for id in batterIDs]))

        #Get contact rate by batter id
        cur.execute('SELECT batter_id, COUNT(CASE WHEN call_id = 2 OR call_id = 3 OR call_id = 4 THEN 1 END), COUNT(CASE WHEN call_id = 3 OR call_id = 4 THEN 1 END) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY batter_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        contact_ratebyID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_discipline SET contact_rate = ? WHERE batter_id = ?', ([(contact_ratebyID.get(id, None), id) for id in batterIDs]))

        #Get inside of zone rate by batter id
        #Get strikes based on universal strike zone
        cur.execute('SELECT batter_id, COUNT(*), COUNT(CASE WHEN location_side >= -0.7508 AND location_side <= 0.7508 AND location_height >= 1.5942 AND location_height <= 3.6033 THEN 1 END) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY batter_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        zone_ratebyID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE batting_stats_discipline SET zone_rate = ? WHERE batter_id = ?', ([(zone_ratebyID.get(id, None), id) for id in batterIDs]))
        
        #Do these on the dashboard side, should be a simple query (player / league)
        # Advanced Plus BB%+, K%+, AVG+, OBP+, SLG+, wRC+, wOBA+, ISO+, BABIP+, xBA+, xSLG+, xwOBA+
        # Batted Ball Plus LD%+, GB%+, FB%+, PULL%+, CENT%+, OPPO%+, PULLGB%+, CENTGB%+, OPPOGB%+, PULLOFFB%+, CENTOFFB%+, OPPOOFFB%+, SOFT%+, MED%+, HARD%+
        # Statcast Plus EV+, maxEV+, Barrel%+, HH%+
        # Discipline Plus O-Swing%+, Z-Swing%+, Swing%+, O-Contact%+, Z-Contact%+, Contact%+, Zone%+
        
        #Pitchers

        #Standard

        #Get player_ids for filling null values and insert every id into the table
        cur.execute('''SELECT trackman.pitcher_id, trackman.league_id, trackman.division_id, pitchers.team_id, teams.year trackman
                    FROM trackman
                    JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
                    JOIN teams ON pitchers.team_id = teams.team_id
                    WHERE (trackman.upload_timestamp + 43200) > ?
                    GROUP BY trackman.pitcher_id, teams.year''', (self.timestamp,))
        tupIDs = [tup + (self.timestamp,) for tup in cur.fetchall()]
        pitcherIDs = [tup[0] for tup in tupIDs]

        #Clear standard table of recent players if they are already listed
        cur.executemany('DELETE FROM pitching_stats_standard WHERE pitcher_id = ?', [(id,) for id in pitcherIDs])
        
        #Insert every player id into the table that played in recently uploaded games
        cur.executemany('INSERT INTO pitching_stats_standard (pitcher_id, league_id, division_id, team_id, year, last_updated) VALUES (?,?,?,?,?,?)', (tupIDs))

        #Get games by pitcher id
        cur.execute('SELECT pitcher_id, COUNT(DISTINCT game_id) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY pitcher_id', (self.timestamp,))
        gByID = dict(cur.fetchall())
        cur.executemany('UPDATE pitching_stats_standard SET g = ? WHERE pitcher_id = ?', ([(gByID.get(id, 0), id) for id in pitcherIDs]))

        #Get games started by pitcher id
        cur.execute('SELECT pitcher_id, COUNT(*) FROM trackman WHERE (upload_timestamp + 43200) > ? AND inning = 1 AND pa_of_inning = 1 AND pitch_of_pa = 1 GROUP BY pitcher_id', (self.timestamp,))
        gsByID = dict(cur.fetchall())
        cur.executemany('UPDATE pitching_stats_standard SET gs = ? WHERE pitcher_id = ?', ([(gsByID.get(id, 0), id) for id in pitcherIDs]))
 
        #Get complete games by pitcher id
        cur.execute('SELECT pitcher_id, COUNT(*) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY game_id, top_bottom_id HAVING COUNT(DISTINCT pitcher_id) = 1', (self.timestamp,))
        cgByID = dict(cur.fetchall())
        cur.executemany('UPDATE pitching_stats_standard SET cg = ? WHERE pitcher_id = ?', ([(cgByID.get(id, 0), id) for id in pitcherIDs]))
 
        #Get shutouts by pitcher id
        cur.execute('SELECT pitcher_id, COUNT(*) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY game_id, top_bottom_id HAVING COUNT(DISTINCT pitcher_id) = 1 AND SUM(runs_scored) = 0', (self.timestamp,))
        shoByID = dict(cur.fetchall())
        cur.executemany('UPDATE pitching_stats_standard SET sho = ? WHERE pitcher_id = ?', ([(shoByID.get(id, 0), id) for id in pitcherIDs]))
 
        #Get innings pitched by pitcher id
        cur.execute('SELECT pitcher_id, COUNT(DISTINCT (game_id || inning || pa_of_inning)) FROM trackman WHERE (result_id = ? OR k_or_bb_id = ?) AND (upload_timestamp + 43200) > ? GROUP BY pitcher_id', (6,2,self.timestamp))
        ipByID = dict(cur.fetchall())
        cur.executemany('UPDATE pitching_stats_standard SET ip = ? WHERE pitcher_id = ?', ([((ipByID.get(id, 0) / 3), id) for id in pitcherIDs]))
 
        #Get total batters faced by pitcher id
        cur.execute('SELECT pitcher_id, COUNT(DISTINCT (game_id || inning || pa_of_inning)) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY pitcher_id', (self.timestamp,))
        tbfByID = dict(cur.fetchall())
        cur.executemany('UPDATE pitching_stats_standard SET tbf = ? WHERE pitcher_id = ?', ([(tbfByID.get(id, 0), id) for id in pitcherIDs]))

        #Get hits by pitcher id
        cur.execute('SELECT pitcher_id, COUNT(*) FROM trackman WHERE (upload_timestamp + 43200) > ? AND result_id <= 4 GROUP BY pitcher_id', (self.timestamp,))
        hByID = dict(cur.fetchall())
        cur.executemany('UPDATE pitching_stats_standard SET h = ? WHERE pitcher_id = ?', ([(hByID.get(id, 0), id) for id in pitcherIDs]))
 
        #Get runs by pitcher id
        cur.execute('SELECT pitcher_id, SUM(runs_scored) FROM trackman WHERE (upload_timestamp + 43200) > ? AND result_id <= 4 GROUP BY pitcher_id', (self.timestamp,))
        rByID = dict(cur.fetchall())
        cur.executemany('UPDATE pitching_stats_standard SET r = ? WHERE pitcher_id = ?', ([(rByID.get(id, 0), id) for id in pitcherIDs]))
 
        #Get home runs by pitcher id
        cur.execute('SELECT pitcher_id, COUNT(*) FROM trackman WHERE (upload_timestamp + 43200) > ? AND result_id = 4 GROUP BY pitcher_id', (self.timestamp,))
        hrByID = dict(cur.fetchall())
        cur.executemany('UPDATE pitching_stats_standard SET hr = ? WHERE pitcher_id = ?', ([(hrByID.get(id, 0), id) for id in pitcherIDs]))

        #Get walks by pitcher id
        cur.execute('SELECT pitcher_id, COUNT(*) FROM trackman WHERE (upload_timestamp + 43200) > ? AND k_or_bb_id = 3 GROUP BY pitcher_id', (self.timestamp,))
        bbByID = dict(cur.fetchall())
        cur.executemany('UPDATE pitching_stats_standard SET bb = ? WHERE pitcher_id = ?', ([(bbByID.get(id, 0), id) for id in pitcherIDs]))

        #Get intentional walks by pitcher id
        cur.execute('SELECT pitcher_id, COUNT(*) FROM trackman WHERE (upload_timestamp + 43200) > ? AND k_or_bb_id = 3 AND call_id = 9 GROUP BY pitcher_id', (self.timestamp,))
        ibbByID = dict(cur.fetchall())
        cur.executemany('UPDATE pitching_stats_standard SET ibb = ? WHERE pitcher_id = ?', ([(ibbByID.get(id, 0), id) for id in pitcherIDs]))
  
        #Get hit by pitch by pitcher id
        cur.execute('SELECT pitcher_id, COUNT(*) FROM trackman WHERE (upload_timestamp + 43200) > ? AND call_id = 6 GROUP BY pitcher_id', (self.timestamp,))
        hbpByID = dict(cur.fetchall())
        cur.executemany('UPDATE pitching_stats_standard SET hbp = ? WHERE pitcher_id = ?', ([(hbpByID.get(id, 0), id) for id in pitcherIDs]))
  
        #Get strikeouts by pitcher id
        cur.execute('SELECT pitcher_id, COUNT(*) FROM trackman WHERE (upload_timestamp + 43200) > ? AND k_or_bb_id = 2 GROUP BY pitcher_id', (self.timestamp,))
        soByID = dict(cur.fetchall())
        cur.executemany('UPDATE pitching_stats_standard SET so = ? WHERE pitcher_id = ?', ([(soByID.get(id, 0), id) for id in pitcherIDs]))
  
        #Get grounded into double plays by pitcher id
        cur.execute('SELECT pitcher_id, COUNT(*) FROM trackman WHERE (upload_timestamp + 43200) > ? AND outs_made >= 2 AND hit_type_id = 4 GROUP BY pitcher_id', (self.timestamp,))
        gdpByID = dict(cur.fetchall())
        cur.executemany('UPDATE pitching_stats_standard SET gdp = ? WHERE pitcher_id = ?', ([(gdpByID.get(id, 0), id) for id in pitcherIDs]))
        
        # Statcast xBA, xSLG, xwOBA

        #Clear statcast table of recent players if they are already listed
        cur.executemany('DELETE FROM pitching_stats_statcast WHERE pitcher_id = ?', [(id,) for id in pitcherIDs])
        
        #Insert every player id into the table that played in recently uploaded games
        cur.executemany('INSERT INTO pitching_stats_statcast (pitcher_id, league_id, division_id, team_id, year, last_updated) VALUES (?,?,?,?,?,?)', (tupIDs))

        #Get every batted ball event by pitcher id
        cur.execute('SELECT pitcher_id, COUNT(*) FROM trackman WHERE (upload_timestamp + 43200) > ? AND call_id = 4 AND exit_velocity > 0 GROUP BY pitcher_id', (self.timestamp,))
        bbeByID = dict(cur.fetchall())
        cur.executemany('UPDATE pitching_stats_statcast SET bbe = ? WHERE pitcher_id = ?', ([(bbeByID.get(id, 0), id) for id in pitcherIDs]))
   
        #Get average exit velo by pitcher id
        cur.execute('SELECT pitcher_id, AVG(exit_velocity) FROM trackman WHERE (upload_timestamp + 43200) > ? AND call_id = 4 AND exit_velocity > 0 GROUP BY pitcher_id', (self.timestamp,))
        avg_evByID = dict(cur.fetchall())
        cur.executemany('UPDATE pitching_stats_statcast SET avg_ev = ? WHERE pitcher_id = ?', ([(avg_evByID.get(id, None), id) for id in pitcherIDs]))
   
        #Get max exit velo by pitcher id
        cur.execute('SELECT pitcher_id, MAX(exit_velocity) FROM trackman WHERE (upload_timestamp + 43200) > ? AND call_id = 4 AND exit_velocity > 0 GROUP BY pitcher_id', (self.timestamp,))
        max_evByID = dict(cur.fetchall())
        cur.executemany('UPDATE pitching_stats_statcast SET max_ev = ? WHERE pitcher_id = ?', ([(max_evByID.get(id, None), id) for id in pitcherIDs]))

        #Get average launch angle by pitcher id
        cur.execute('SELECT pitcher_id, AVG(launch_angle) FROM trackman WHERE (upload_timestamp + 43200) > ? AND call_id = 4 AND exit_velocity > 0 GROUP BY pitcher_id', (self.timestamp,))
        avg_laByID = dict(cur.fetchall())
        cur.executemany('UPDATE pitching_stats_statcast SET avg_la = ? WHERE pitcher_id = ?', ([(avg_laByID.get(id, None), id) for id in pitcherIDs]))
 
        #Get barrels by pitcher id (probably need a league specific definition of barrel) 
        cur.execute('SELECT pitcher_id, exit_velocity, launch_angle FROM trackman WHERE (upload_timestamp + 43200) > ? AND call_id = 4 AND exit_velocity > 0', (self.timestamp,))
        tupsByID = cur.fetchall()

        brlsByID = Counter()
        for tup in tupsByID:
            id = tup[0]
            exit_velocity = float(tup[1])
            launch_angle = float(tup[2])
            barrel = False
            if exit_velocity >= 98.0:
                if exit_velocity <= 99.0:
                    if launch_angle >= 26.0 and launch_angle <= 30.0:
                        barrel = True
                elif exit_velocity <= 100.0:
                    if launch_angle >= 25.0 and launch_angle <= 31.0:
                        barrel = True
                #Not a perfect representation of a barrel, but pretty close
                else:
                    range_growth = (exit_velocity - 100.0) * 1.2
                    high_angle = min(31.0 + range_growth, 50.0)
                    low_angle = max(25.0 - range_growth, 8.0)
                    if launch_angle >= low_angle and launch_angle <= high_angle:
                        barrel = True
            if barrel:
                brlsByID[id] += 1
        brlsByID = dict(brlsByID)
        cur.executemany('UPDATE pitching_stats_statcast SET brls = ? WHERE pitcher_id = ?', ([(brlsByID.get(id, 0), id) for id in pitcherIDs]))
        
        #Get barrel rate by pitcher id
        cur.execute('SELECT pitcher_id, bbe, brls FROM pitching_stats_statcast WHERE (last_updated + 43200) > ? ', (self.timestamp,))
        tupsByID = cur.fetchall()
        brl_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_statcast SET brl_rate = ? WHERE pitcher_id = ?', ([(brl_rateByID.get(id, None), id) for id in pitcherIDs]))
        
        #Get hard hits by pitcher id
        cur.execute('SELECT pitcher_id, COUNT(*) FROM trackman WHERE (upload_timestamp + 43200) > ? AND call_id = 4 AND exit_velocity >= 95 GROUP BY pitcher_id', (self.timestamp,))
        hhByID = dict(cur.fetchall())
        cur.executemany('UPDATE pitching_stats_statcast SET hh = ? WHERE pitcher_id = ?', ([(hhByID.get(id, 0), id) for id in pitcherIDs]))
        
        #Get hard hit rate by pitcher id
        cur.execute('SELECT pitcher_id, bbe, hh FROM pitching_stats_statcast WHERE (last_updated + 43200) > ?', (self.timestamp,))
        tupsByID = cur.fetchall()
        hh_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_statcast SET hh_rate = ? WHERE pitcher_id = ?', ([(hh_rateByID.get(id, None), id) for id in pitcherIDs]))
        
        # Advanced K-BB%,FIP, SIERA

         #Clear advanced table of recent players if they are already listed
        cur.executemany('DELETE FROM pitching_stats_advanced WHERE pitcher_id = ?', [(id,) for id in pitcherIDs])
        
        #Insert every player id into the table that played in recently uploaded games
        cur.executemany('INSERT INTO pitching_stats_advanced (pitcher_id, league_id, division_id, team_id, year, last_updated) VALUES (?,?,?,?,?,?)', (tupIDs))

        #Get r/9 by pitcher id
        cur.execute('SELECT pitcher_id, ip, r FROM pitching_stats_standard WHERE (last_updated + 43200) > ?', (self.timestamp,))
        tupsByID = cur.fetchall()
        r_9ByID = {tup[0] : (tup[2] / tup[1]) * 9 for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_advanced SET r_9 = ? WHERE pitcher_id = ?', ([(r_9ByID.get(id, None), id) for id in pitcherIDs]))

        #Get k/9 by pitcher id
        cur.execute('SELECT pitcher_id, ip, so FROM pitching_stats_standard WHERE (last_updated + 43200) > ?', (self.timestamp,))
        tupsByID = cur.fetchall()
        k_9ByID = {tup[0] : (tup[2] / tup[1]) * 9 for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_advanced SET k_9 = ? WHERE pitcher_id = ?', ([(k_9ByID.get(id, None), id) for id in pitcherIDs]))

        #Get bb/9 by pitcher id
        cur.execute('SELECT pitcher_id, ip, bb FROM pitching_stats_standard WHERE (last_updated + 43200) > ?', (self.timestamp,))
        tupsByID = cur.fetchall()
        bb_9ByID = {tup[0] : (tup[2] / tup[1]) * 9 for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_advanced SET bb_9 = ? WHERE pitcher_id = ?', ([(bb_9ByID.get(id, None), id) for id in pitcherIDs]))

        #Get k/bb by pitcher id
        cur.execute('SELECT pitcher_id, so, bb FROM pitching_stats_standard WHERE (last_updated + 43200) > ?', (self.timestamp,))
        tupsByID = cur.fetchall()
        k_bbByID = {tup[0] : tup[1] / tup[2] for tup in tupsByID if tup[2]}
        cur.executemany('UPDATE pitching_stats_advanced SET k_bb = ? WHERE pitcher_id = ?', ([(k_bbByID.get(id, None), id) for id in pitcherIDs]))

        #Get hr/9 by pitcher id
        cur.execute('SELECT pitcher_id, ip, hr FROM pitching_stats_standard WHERE (last_updated + 43200) > ?', (self.timestamp,))
        tupsByID = cur.fetchall()
        hr_9ByID = {tup[0] : (tup[2] / tup[1]) * 9 for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_advanced SET hr_9 = ? WHERE pitcher_id = ?', ([(hr_9ByID.get(id, None), id) for id in pitcherIDs]))

        #Get strikeout rate by pitcher id
        cur.execute('SELECT pitcher_id, tbf, so FROM pitching_stats_standard WHERE (last_updated + 43200) > ?', (self.timestamp,))
        tupsByID = cur.fetchall()
        so_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_advanced SET k_rate = ? WHERE pitcher_id = ?', ([(so_rateByID.get(id, None), id) for id in pitcherIDs]))

        #Get walk rate by pitcher id
        cur.execute('SELECT pitcher_id, tbf, bb FROM pitching_stats_standard WHERE (last_updated + 43200) > ?', (self.timestamp,))
        tupsByID = cur.fetchall()
        bb_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_advanced SET bb_rate = ? WHERE pitcher_id = ?', ([(bb_rateByID.get(id, None), id) for id in pitcherIDs]))

        #Get whip by pitcher id
        cur.execute('SELECT pitcher_id, ip, bb, h FROM pitching_stats_standard WHERE (last_updated + 43200) > ?', (self.timestamp,))
        tupsByID = cur.fetchall()
        whipByID = {tup[0] : (tup[2] + tup[3]) / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_advanced SET whip = ? WHERE pitcher_id = ?', ([(whipByID.get(id, None), id) for id in pitcherIDs]))
 
        #Get babip by pitcher id
        cur.execute('SELECT pitcher_id, COUNT(*), COUNT(CASE WHEN result_id <= 4 THEN 1 END) FROM trackman WHERE call_id = 4 AND (upload_timestamp + 43200) > ? GROUP BY pitcher_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        babipByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_advanced SET babip = ? WHERE pitcher_id = ?', ([(babipByID.get(id, None), id) for id in pitcherIDs]))
        
        # Batted Ball 

         #Clear batted ball table of recent players if they are already listed
        cur.executemany('DELETE FROM pitching_stats_batted_ball WHERE pitcher_id = ?', [(id,) for id in pitcherIDs])
        
        #Insert every player id into the table that played in recently uploaded games
        cur.executemany('INSERT INTO pitching_stats_batted_ball (pitcher_id, league_id, division_id, team_id, year, last_updated) VALUES (?,?,?,?,?,?)', (tupIDs))

        #Insert every batted ball event by pitcher id
        cur.executemany('UPDATE pitching_stats_batted_ball SET bbe = ? WHERE pitcher_id = ?', ([(bbeByID.get(id, 0), id) for id in pitcherIDs]))
        
        #Get every ground ball rate by pitcher id
        cur.execute('SELECT pitcher_id, COUNT(*), COUNT(CASE WHEN launch_angle <= 10 THEN 1 END) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY pitcher_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        gb_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_batted_ball SET gb_rate = ? WHERE pitcher_id = ?', ([(gb_rateByID.get(id, None), id) for id in pitcherIDs]))

        #Get every fly ball rate by pitcher id 
        cur.execute('SELECT pitcher_id, COUNT(*), COUNT(CASE WHEN launch_angle > 25 AND launch_angle <=50 THEN 1 END) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY pitcher_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        fb_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_batted_ball SET fb_rate = ? WHERE pitcher_id = ?', ([(fb_rateByID.get(id, None), id) for id in pitcherIDs]))

        #Get every line drive rate by pitcher id
        cur.execute('SELECT pitcher_id, COUNT(*), COUNT(CASE WHEN launch_angle > 10 AND launch_angle <=25 THEN 1 END) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY pitcher_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        ld_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_batted_ball SET ld_rate = ? WHERE pitcher_id = ?', ([(ld_rateByID.get(id, None), id) for id in pitcherIDs]))

        #Get every infield fly ball rate rate by pitcher id
        cur.execute('SELECT pitcher_id, COUNT(*), COUNT(CASE WHEN launch_angle > 50 THEN 1 END) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY pitcher_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        iffb_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_batted_ball SET iffb_rate = ? WHERE pitcher_id = ?', ([(iffb_rateByID.get(id, None), id) for id in pitcherIDs]))

        #Get every home run per fly ball rate by pitcher id
        cur.execute('SELECT pitcher_id, COUNT(CASE WHEN launch_angle > 25 AND launch_angle <=50 THEN 1 END), COUNT(CASE WHEN launch_angle > 25 AND launch_angle <=50 AND result_id = 4 THEN 1 END) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY pitcher_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        hr_fbByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_batted_ball SET hr_fb = ? WHERE pitcher_id = ?', ([(hr_fbByID.get(id, None), id) for id in pitcherIDs]))

        #Get every strike rate rate by pitcher id
        cur.execute('SELECT pitcher_id, COUNT(*), COUNT(CASE WHEN call_id <= 4 THEN 1 END) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY pitcher_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        strike_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_batted_ball SET strike_rate = ? WHERE pitcher_id = ?', ([(strike_rateByID.get(id, None), id) for id in pitcherIDs]))

        #Get every ball rate rate by pitcher id
        cur.execute('SELECT pitcher_id, COUNT(*), COUNT(CASE WHEN call_id >= 5 AND call_id <= 7 THEN 1 END) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY pitcher_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        ball_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_batted_ball SET ball_rate = ? WHERE pitcher_id = ?', ([(ball_rateByID.get(id, None), id) for id in pitcherIDs]))

        #Get every pull rate by pitcher id
        cur.execute('''SELECT trackman.pitcher_id, COUNT(*),
        COUNT(CASE WHEN (trackman.hit_bearing < -15 AND batters.batter_side_id = 1) OR (trackman.hit_bearing > 15 AND batters.batter_side_id = 2) OR
        (trackman.hit_bearing > 15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 1) OR
        (trackman.hit_bearing < -15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 2) THEN 1 END)
        FROM trackman
        LEFT JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
        LEFT JOIN batters ON trackman.batter_id = batters.batter_id
        WHERE trackman.call_id = 4 AND trackman.exit_velocity > 0 AND trackman.hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ? GROUP BY trackman.pitcher_id''', (self.timestamp,))
        tupsByID = cur.fetchall()
        pull_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_batted_ball SET pull_rate = ? WHERE pitcher_id = ?', ([(pull_rateByID.get(id, None), id) for id in pitcherIDs]))

        #Get every center rate by pitcher id
        cur.execute('''SELECT pitcher_id, COUNT(*),
        COUNT(CASE WHEN hit_bearing >= -15 AND hit_bearing <= 15 THEN 1 END)
        FROM trackman
        WHERE call_id = 4 AND exit_velocity > 0 AND hit_bearing IS NOT NULL AND (upload_timestamp + 43200) > ? GROUP BY pitcher_id''', (self.timestamp,))
        tupsByID = cur.fetchall()
        cent_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_batted_ball SET cent_rate = ? WHERE pitcher_id = ?', ([(cent_rateByID.get(id, None), id) for id in pitcherIDs]))
 
        #Get every oppo rate by pitcher id
        cur.execute('''SELECT trackman.pitcher_id, COUNT(*),
        COUNT(CASE WHEN (trackman.hit_bearing > 15 AND batters.batter_side_id = 1) OR (trackman.hit_bearing < -15 AND batters.batter_side_id = 2) OR
        (trackman.hit_bearing < -15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 1) OR
        (trackman.hit_bearing > 15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 2) THEN 1 END)
        FROM trackman
        LEFT JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
        LEFT JOIN batters ON trackman.batter_id = batters.batter_id
        WHERE trackman.call_id = 4 AND trackman.exit_velocity > 0 AND trackman.hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ? GROUP BY trackman.pitcher_id''', (self.timestamp,))
        tupsByID = cur.fetchall()
        oppo_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_batted_ball SET oppo_rate = ? WHERE pitcher_id = ?', ([(oppo_rateByID.get(id, None), id) for id in pitcherIDs]))

        #Get every pull per gb rate by pitcher id 
        cur.execute('''SELECT trackman.pitcher_id, COUNT(*),
        COUNT(CASE WHEN (trackman.hit_bearing < -15 AND batters.batter_side_id = 1) OR (trackman.hit_bearing > 15 AND batters.batter_side_id = 2) OR
        (trackman.hit_bearing > 15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 1) OR
        (trackman.hit_bearing < -15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 2) THEN 1 END)
        FROM trackman
        LEFT JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
        LEFT JOIN batters ON trackman.batter_id = batters.batter_id
        WHERE trackman.call_id = 4 AND trackman.exit_velocity > 0  AND trackman.launch_angle <= 10 AND trackman.hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ? GROUP BY trackman.pitcher_id''', (self.timestamp,))
        tupsByID = cur.fetchall()
        pull_gbByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_batted_ball SET pull_gb_rate = ? WHERE pitcher_id = ?', ([(pull_gbByID.get(id, None), id) for id in pitcherIDs]))

        #Get every center per gb rate by pitcher id
        cur.execute('''SELECT pitcher_id, COUNT(*),
        COUNT(CASE WHEN hit_bearing >= -15 AND hit_bearing <= 15 THEN 1 END)
        FROM trackman
        WHERE call_id = 4 AND exit_velocity > 0 AND launch_angle <= 10 AND hit_bearing IS NOT NULL AND (upload_timestamp + 43200) > ? GROUP BY pitcher_id''', (self.timestamp,))
        tupsByID = cur.fetchall()
        cent_gbByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_batted_ball SET cent_gb_rate = ? WHERE pitcher_id = ?', ([(cent_gbByID.get(id, None), id) for id in pitcherIDs]))

        #Get every oppo per gb rate by pitcher id
        cur.execute('''SELECT trackman.pitcher_id, COUNT(*),
        COUNT(CASE WHEN (trackman.hit_bearing > 15 AND batters.batter_side_id = 1) OR (trackman.hit_bearing < -15 AND batters.batter_side_id = 2) OR
        (trackman.hit_bearing < -15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 1) OR
        (trackman.hit_bearing > 15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 2) THEN 1 END)
        FROM trackman
        LEFT JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
        LEFT JOIN batters ON trackman.batter_id = batters.batter_id
        WHERE trackman.call_id = 4 AND trackman.exit_velocity > 0 AND trackman.launch_angle <= 10 AND trackman.hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ? GROUP BY trackman.pitcher_id''', (self.timestamp,))
        tupsByID = cur.fetchall()
        oppo_gbByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_batted_ball SET oppo_gb_rate = ? WHERE pitcher_id = ?', ([(oppo_gbByID.get(id, None), id) for id in pitcherIDs]))

        #Get every pull per fb rate by pitcher id (use lower, lower bound for launch angle for purpose of stat--to be used for outfield positioning)
        cur.execute('''SELECT trackman.pitcher_id, COUNT(*),
        COUNT(CASE WHEN (trackman.hit_bearing < -15 AND batters.batter_side_id = 1) OR (trackman.hit_bearing > 15 AND batters.batter_side_id = 2) OR
        (trackman.hit_bearing > 15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 1) OR
        (trackman.hit_bearing < -15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 2) THEN 1 END)
        FROM trackman
        LEFT JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
        LEFT JOIN batters ON trackman.batter_id = batters.batter_id
        WHERE trackman.call_id = 4 AND trackman.exit_velocity > 0  AND trackman.launch_angle > 20 AND trackman.launch_angle <=50 AND trackman.hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ? GROUP BY trackman.pitcher_id''', (self.timestamp,))
        tupsByID = cur.fetchall()
        pull_fbByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_batted_ball SET pull_fb_rate = ? WHERE pitcher_id = ?', ([(pull_fbByID.get(id, None), id) for id in pitcherIDs]))

        #Get every center per fb rate by pitcher id (use lower, lower bound for launch angle for purpose of stat--to be used for outfield positioning)
        cur.execute('''SELECT pitcher_id, COUNT(*),
        COUNT(CASE WHEN hit_bearing >= -15 AND hit_bearing <= 15 THEN 1 END)
        FROM trackman
        WHERE call_id = 4 AND exit_velocity > 0 AND launch_angle > 20 AND launch_angle <=50 AND hit_bearing IS NOT NULL AND (upload_timestamp + 43200) > ? GROUP BY pitcher_id''', (self.timestamp,))
        tupsByID = cur.fetchall()
        cent_fbByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_batted_ball SET cent_fb_rate = ? WHERE pitcher_id = ?', ([(cent_fbByID.get(id, None), id) for id in pitcherIDs]))

        #Get every oppo per fb rate by pitcher id (use lower, lower bound for launch angle for purpose of stat--to be used for outfield positioning)
        cur.execute('''SELECT trackman.pitcher_id, COUNT(*),
        COUNT(CASE WHEN (trackman.hit_bearing > 15 AND batters.batter_side_id = 1) OR (trackman.hit_bearing < -15 AND batters.batter_side_id = 2) OR
        (trackman.hit_bearing < -15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 1) OR
        (trackman.hit_bearing > 15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 2) THEN 1 END)
        FROM trackman
        LEFT JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
        LEFT JOIN batters ON trackman.batter_id = batters.batter_id
        WHERE trackman.call_id = 4 AND trackman.exit_velocity > 0 AND trackman.launch_angle > 20 AND trackman.launch_angle <=50 AND trackman.hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ? GROUP BY trackman.pitcher_id''', (self.timestamp,))
        tupsByID = cur.fetchall()
        oppo_fbByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_batted_ball SET oppo_fb_rate = ? WHERE pitcher_id = ?', ([(oppo_fbByID.get(id, None), id) for id in pitcherIDs]))

        #Get soft con rate by pitcher id
        cur.execute('SELECT pitcher_id, COUNT(*), COUNT(CASE WHEN exit_velocity < 70 THEN 1 END) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY pitcher_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        soft_conbyID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_batted_ball SET soft_con_rate = ? WHERE pitcher_id = ?', ([(soft_conbyID.get(id, None), id) for id in pitcherIDs]))

        #Get med con rate by pitcher id
        cur.execute('SELECT pitcher_id, COUNT(*), COUNT(CASE WHEN exit_velocity >= 70 AND exit_velocity < 95 THEN 1 END) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY pitcher_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        med_conbyID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_batted_ball SET med_con_rate = ? WHERE pitcher_id = ?', ([(med_conbyID.get(id, None), id) for id in pitcherIDs]))

        #Get hard con rate by pitcher id
        cur.execute('SELECT pitcher_id, COUNT(*), COUNT(CASE WHEN exit_velocity >= 95 THEN 1 END) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY pitcher_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        hard_conbyID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_batted_ball SET hard_con_rate = ? WHERE pitcher_id = ?', ([(hard_conbyID.get(id, None), id) for id in pitcherIDs]))
        
        # Discipline 

         #Clear discipline table of recent players if they are already listed
        cur.executemany('DELETE FROM pitching_stats_discipline WHERE pitcher_id = ?', [(id,) for id in pitcherIDs])
        
        #Insert every player id into the table that played in recently uploaded games
        cur.executemany('INSERT INTO pitching_stats_discipline (pitcher_id, league_id, division_id, team_id, year, last_updated) VALUES (?,?,?,?,?,?)', (tupIDs))

        #Get outside of zone swing rate by pitcher id
        #Get balls based on universal strike zone
        cur.execute('SELECT pitcher_id, COUNT(*), COUNT(CASE WHEN call_id = 2 OR call_id = 3 OR call_id = 4 THEN 1 END) FROM trackman WHERE (upload_timestamp + 43200) > ? AND location_side < -0.7508 OR location_side > 0.7508 OR location_height < 1.5942 OR location_height > 3.6033 GROUP BY pitcher_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        o_swingbyID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_discipline SET o_swing = ? WHERE pitcher_id = ?', ([(o_swingbyID.get(id, None), id) for id in pitcherIDs]))

        #Get inside of zone swing rate by pitcher id
        #Get strikes based on universal strike zone
        cur.execute('SELECT pitcher_id, COUNT(*), COUNT(CASE WHEN call_id = 2 OR call_id = 3 OR call_id = 4 THEN 1 END) FROM trackman WHERE (upload_timestamp + 43200) > ? AND location_side >= -0.7508 AND location_side <= 0.7508 AND location_height >= 1.5942 AND location_height <= 3.6033 GROUP BY pitcher_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        z_swingbyID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_discipline SET z_swing = ? WHERE pitcher_id = ?', ([(z_swingbyID.get(id, None), id) for id in pitcherIDs]))

        #Get swing rate by pitcher id
        cur.execute('SELECT pitcher_id, COUNT(*), COUNT(CASE WHEN call_id = 2 OR call_id = 3 OR call_id = 4 THEN 1 END) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY pitcher_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        swing_ratebyID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_discipline SET swing_rate = ? WHERE pitcher_id = ?', ([(swing_ratebyID.get(id, None), id) for id in pitcherIDs]))

        #Get outside of zone contact rate by pitcher id
        #Get balls based on universal strike zone
        cur.execute('SELECT pitcher_id, COUNT(CASE WHEN call_id = 2 OR call_id = 3 OR call_id = 4 THEN 1 END), COUNT(CASE WHEN call_id = 3 OR call_id = 4 THEN 1 END) FROM trackman WHERE (upload_timestamp + 43200) > ? AND location_side < -0.7508 OR location_side > 0.7508 OR location_height < 1.5942 OR location_height > 3.6033 GROUP BY pitcher_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        o_contactbyID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_discipline SET o_contact = ? WHERE pitcher_id = ?', ([(o_contactbyID.get(id, None), id) for id in pitcherIDs]))

        #Get inside of zone contact rate by pitcher id
        #Get strikes based on universal strike zone
        cur.execute('SELECT pitcher_id, COUNT(CASE WHEN call_id = 2 OR call_id = 3 OR call_id = 4 THEN 1 END), COUNT(CASE WHEN call_id = 3 OR call_id = 4 THEN 1 END) FROM trackman WHERE (upload_timestamp + 43200) > ? AND location_side >= -0.7508 AND location_side <= 0.7508 AND location_height >= 1.5942 AND location_height <= 3.6033 GROUP BY pitcher_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        z_contactbyID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_discipline SET z_contact = ? WHERE pitcher_id = ?', ([(z_contactbyID.get(id, None), id) for id in pitcherIDs]))

        #Get contact rate by pitcher id
        cur.execute('SELECT pitcher_id, COUNT(CASE WHEN call_id = 2 OR call_id = 3 OR call_id = 4 THEN 1 END), COUNT(CASE WHEN call_id = 3 OR call_id = 4 THEN 1 END) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY pitcher_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        contact_ratebyID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_discipline SET contact_rate = ? WHERE pitcher_id = ?', ([(contact_ratebyID.get(id, None), id) for id in pitcherIDs]))

        #Get inside of zone rate by pitcher id
        #Get strikes based on universal strike zone
        cur.execute('SELECT pitcher_id, COUNT(*), COUNT(CASE WHEN location_side >= -0.7508 AND location_side <= 0.7508 AND location_height >= 1.5942 AND location_height <= 3.6033 THEN 1 END) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY pitcher_id', (self.timestamp,))
        tupsByID = cur.fetchall()
        zone_ratebyID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE pitching_stats_discipline SET zone_rate = ? WHERE pitcher_id = ?', ([(zone_ratebyID.get(id, None), id) for id in pitcherIDs]))
        
        #Do these on the dashboard side, should be a simple query (player / league)
        # Advanced Plus K/9+, BB/9+, K/BB+, HR/9+, K%+, BB%+, AVG+, WHIP+, BABIP+, ERA-, FIP-, xFIP-, LD%+, GB%+, FB%+, PULL%+, CENT%+, OPPO%+, PULLGB%+, CENTGB%+, OPPOGB%+, PULLOFFB%+, CENTOFFB%+, OPPOOFFB%+, SOFT%+, MED%+, HARD%+
        # Statcast Plus EV+, maxEV+, Barrel%+, HH%+, xBA+, xSLG+, xwOBA+
        # Discipline Plus O-Swing%+, Z-Swing%+, Swing%+, O-Contact%+, Z-Contact%+, Contact%+, Zone%+
        
        #Pitcher Arsenals

        # Standard

        #Get pitch ids from tagged pitch types for filling null values and insert every id into the table
        cur.execute('''SELECT DISTINCT '1' || trackman.pitcher_id || trackman.tagged_type_id as tagged_pitch_id, 
                    trackman.pitcher_id, trackman.league_id, trackman.division_id, pitchers.team_id, teams.year, trackman.tagged_type_id
                    FROM trackman
                    JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
                    JOIN teams ON pitchers.team_id = teams.team_id
                    WHERE (trackman.upload_timestamp + 43200) > ?
                    GROUP BY tagged_pitch_id

                    UNION

                    SELECT DISTINCT '2' || trackman.pitcher_id || trackman.auto_type_id as auto_pitch_id,
                    trackman.pitcher_id, trackman.league_id, trackman.division_id, pitchers.team_id, teams.year, trackman.auto_type_id 
                    FROM trackman
                    JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
                    JOIN teams ON pitchers.team_id = teams.team_id
                    WHERE (trackman.upload_timestamp + 43200) > ?
                    GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        tupIDs = [tup + (self.timestamp,) for tup in cur.fetchall()]
        arsenalIDs = [tup[0] for tup in tupIDs]

        #Clear standard table of recent players if they are already listed
        cur.executemany('DELETE FROM arsenal_stats_standard WHERE pitch_id = ?', [(id,) for id in arsenalIDs])
        
        #Insert every pitch id into the table that played in recently uploaded games
        cur.executemany('INSERT INTO arsenal_stats_standard (pitch_id, pitcher_id, league_id, division_id, team_id, year, type_id, last_updated) VALUES (?,?,?,?,?,?,?,?)', (tupIDs))

        #Get hits by arsenal id
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, COUNT(*) FROM trackman WHERE result_id <= 4 AND (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, COUNT(*) FROM trackman WHERE result_id <= 4 AND (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        hByID = dict(cur.fetchall())
        cur.executemany('UPDATE arsenal_stats_standard SET h = ? WHERE pitch_id = ?', ([(hByID.get(id, 0), id) for id in arsenalIDs]))
  
        #Get home runs by arsenal id
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, COUNT(*) FROM trackman WHERE result_id = 4 AND (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, COUNT(*) FROM trackman WHERE result_id = 4 AND (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        hrByID = dict(cur.fetchall())
        cur.executemany('UPDATE arsenal_stats_standard SET hr = ? WHERE pitch_id = ?', ([(hrByID.get(id, 0), id) for id in arsenalIDs]))
   
        #Get every strike rate by arsenal id
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, COUNT(*), COUNT(CASE WHEN call_id <= 4 THEN 1 END) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, COUNT(*), COUNT(CASE WHEN call_id <= 4 THEN 1 END) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        tupsByID = cur.fetchall()
        strike_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_standard SET strike_rate = ? WHERE pitch_id = ?', ([(strike_rateByID.get(id, None), id) for id in arsenalIDs]))

        #Get every ball rate by arsenal id
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, COUNT(*), COUNT(CASE WHEN call_id >= 5 AND call_id <= 7 THEN 1 END) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, COUNT(*), COUNT(CASE WHEN call_id >= 5 AND call_id <= 7 THEN 1 END) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        tupsByID = cur.fetchall()
        ball_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_standard SET ball_rate = ? WHERE pitch_id = ?', ([(ball_rateByID.get(id, None), id) for id in arsenalIDs]))

        #Get every number of pitches by arsenal id
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, COUNT(*) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, COUNT(*) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        npByID = dict(cur.fetchall())
        cur.executemany('UPDATE arsenal_stats_standard SET np = ? WHERE pitch_id = ?', ([(npByID.get(id, 0), id) for id in arsenalIDs]))
 
        #Get every number of pitches by aresenal id
        cur.execute('''SELECT pitcher_id, COUNT(*) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY pitcher_id''', (self.timestamp,))
        pitcher_npByID = dict(cur.fetchall())
        cur.executemany('UPDATE arsenal_stats_standard SET pitcher_np = ? WHERE pitcher_id = ?', ([(pitcher_npByID.get(id, 0), id) for id in pitcherIDs]))
        
        #Get every usage rate by arsenal id
        cur.execute('''SELECT pitch_id, pitcher_np, np FROM arsenal_stats_standard WHERE (last_updated + 43200) > ?''', (self.timestamp,))
        tupsByID = cur.fetchall()
        pitch_usageByID = {str(tup[0]) : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_standard SET pitch_usage = ? WHERE pitch_id = ?', ([(pitch_usageByID.get(id, 0), id) for id in arsenalIDs]))
        
        #Get every hit per np rate by arsenal id
        cur.execute('''SELECT pitch_id, np, h FROM arsenal_stats_standard WHERE (last_updated + 43200) > ?''', (self.timestamp,))
        tupsByID = cur.fetchall()
        h_npByID = {str(tup[0]) : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_standard SET h_np = ? WHERE pitch_id = ?', ([(h_npByID.get(id, 0), id) for id in arsenalIDs]))
        
        #Get every home run per np rate by arsenal id
        cur.execute('''SELECT pitch_id, np, hr FROM arsenal_stats_standard WHERE (last_updated + 43200) > ?''', (self.timestamp,))
        tupsByID = cur.fetchall()
        hr_npByID = {str(tup[0]) : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_standard SET hr_np = ? WHERE pitch_id = ?', ([(hr_npByID.get(id, 0), id) for id in arsenalIDs]))
        
        #Get runs by arsenal id
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, SUM(runs_scored) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, SUM(runs_scored) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        rByID = dict(cur.fetchall())
        cur.executemany('UPDATE arsenal_stats_standard SET r = ? WHERE pitch_id = ?', ([(rByID.get(id, 0), id) for id in arsenalIDs]))
        
        #Get every run per np rate by arsenal id
        cur.execute('''SELECT pitch_id, np, r FROM arsenal_stats_standard WHERE (last_updated + 43200) > ?''', (self.timestamp,))
        tupsByID = cur.fetchall()
        r_npByID = {str(tup[0]) : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_standard SET r_np = ? WHERE pitch_id = ?', ([(r_npByID.get(id, 0), id) for id in arsenalIDs]))
        
        # Arsenal Info

        #Clear info table of recent players if they are already listed
        cur.executemany('DELETE FROM arsenal_stats_info WHERE pitch_id = ?', [(id,) for id in arsenalIDs])
        
        #Insert every pitch id into the table that played in recently uploaded games
        cur.executemany('INSERT INTO arsenal_stats_info (pitch_id, pitcher_id, league_id, division_id, team_id, year, type_id, last_updated) VALUES (?,?,?,?,?,?,?,?)', (tupIDs))

        #Get avg velo by pitch id
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, AVG(velocity) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, AVG(velocity) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        avg_veloByID = dict(cur.fetchall())
        cur.executemany('UPDATE arsenal_stats_info SET avg_velo = ? WHERE pitch_id = ?', ([(avg_veloByID.get(id, 0), id) for id in arsenalIDs]))

        #Get max velo by pitch id
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, MAX(velocity) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, MAX(velocity) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        max_veloByID = dict(cur.fetchall())
        cur.executemany('UPDATE arsenal_stats_info SET max_velo = ? WHERE pitch_id = ?', ([(max_veloByID.get(id, 0), id) for id in arsenalIDs]))

        #Get avg spin by pitch id
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, AVG(spin) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, AVG(spin) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        avg_spin_rateByID = dict(cur.fetchall())
        cur.executemany('UPDATE arsenal_stats_info SET avg_spin_rate = ? WHERE pitch_id = ?', ([(avg_spin_rateByID.get(id, 0), id) for id in arsenalIDs]))

        #Get max spin by pitch id
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, MAX(spin) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, MAX(spin) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        max_spin_rateByID = dict(cur.fetchall())
        cur.executemany('UPDATE arsenal_stats_info SET max_spin_rate = ? WHERE pitch_id = ?', ([(max_spin_rateByID.get(id, 0), id) for id in arsenalIDs]))

        #Get avg vert movement by pitch id
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, AVG(vertical) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, AVG(vertical) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        avg_vert_movementByID = dict(cur.fetchall())
        cur.executemany('UPDATE arsenal_stats_info SET avg_vert_movement = ? WHERE pitch_id = ?', ([(avg_vert_movementByID.get(id, 0), id) for id in arsenalIDs]))

        #Get avg induced movement by pitch id
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, AVG(induced) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, AVG(induced) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        avg_induced_movementByID = dict(cur.fetchall())
        cur.executemany('UPDATE arsenal_stats_info SET avg_induced_movement = ? WHERE pitch_id = ?', ([(avg_induced_movementByID.get(id, 0), id) for id in arsenalIDs]))
 
        #Get avg horz movement by pitch id
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, AVG(horizontal) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, AVG(horizontal) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        avg_horz_movementByID = dict(cur.fetchall())
        cur.executemany('UPDATE arsenal_stats_info SET avg_horz_movement = ? WHERE pitch_id = ?', ([(avg_horz_movementByID.get(id, 0), id) for id in arsenalIDs]))
        
        # Arsenal Statcast xBABIP

        #Clear statcast table of recent players if they are already listed
        cur.executemany('DELETE FROM arsenal_stats_statcast WHERE pitch_id = ?', [(id,) for id in arsenalIDs])
        
        #Insert every pitch id into the table that played in recently uploaded games
        cur.executemany('INSERT INTO arsenal_stats_statcast (pitch_id, pitcher_id, league_id, division_id, team_id, year, type_id, last_updated) VALUES (?,?,?,?,?,?,?,?)', (tupIDs))

        #Get avg exit velo by pitch id
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, AVG(exit_velocity) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, AVG(exit_velocity) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        avg_evByID = dict(cur.fetchall())
        cur.executemany('UPDATE arsenal_stats_statcast SET avg_ev = ? WHERE pitch_id = ?', ([(avg_evByID.get(id, 0), id) for id in arsenalIDs]))
 
        #Get max exit velo by pitch id
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, MAX(exit_velocity) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, MAX(exit_velocity) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        max_evByID = dict(cur.fetchall())
        cur.executemany('UPDATE arsenal_stats_statcast SET max_ev = ? WHERE pitch_id = ?', ([(max_evByID.get(id, 0), id) for id in arsenalIDs]))
   
        #Get avg launch angle by pitch id
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, AVG(launch_angle) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, AVG(launch_angle) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        avg_laByID = dict(cur.fetchall())
        cur.executemany('UPDATE arsenal_stats_statcast SET avg_la = ? WHERE pitch_id = ?', ([(avg_laByID.get(id, 0), id) for id in arsenalIDs]))

        #Get every batted ball event by pitch id
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, COUNT(*) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, COUNT(*) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        bbeByID = dict(cur.fetchall())
        cur.executemany('UPDATE arsenal_stats_statcast SET bbe = ? WHERE pitch_id = ?', ([(bbeByID.get(id, 0), id) for id in arsenalIDs]))

        #Get barrels by pitch id (probably need a league specific definition of barrel) 
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, exit_velocity, launch_angle FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ?
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, exit_velocity, launch_angle FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ?''', (self.timestamp,self.timestamp))
        tupsByID = cur.fetchall()

        brlsByID = Counter()
        for tup in tupsByID:
            id = tup[0]
            exit_velocity = float(tup[1])
            launch_angle = float(tup[2])
            barrel = False
            if exit_velocity >= 98.0:
                if exit_velocity <= 99.0:
                    if launch_angle >= 26.0 and launch_angle <= 30.0:
                        barrel = True
                elif exit_velocity <= 100.0:
                    if launch_angle >= 25.0 and launch_angle <= 31.0:
                        barrel = True
                #Not a perfect representation of a barrel, but pretty close
                else:
                    range_growth = (exit_velocity - 100.0) * 1.2
                    high_angle = min(31.0 + range_growth, 50.0)
                    low_angle = max(25.0 - range_growth, 8.0)
                    if launch_angle >= low_angle and launch_angle <= high_angle:
                        barrel = True
            if barrel:
                brlsByID[id] += 1
        brlsByID = dict(brlsByID)
        cur.executemany('UPDATE arsenal_stats_statcast SET brls = ? WHERE pitch_id = ?', ([(brlsByID.get(id, 0), id) for id in arsenalIDs]))
        
        #Get barrel rate by pitch id
        cur.execute('SELECT pitch_id, bbe, brls FROM arsenal_stats_statcast WHERE (last_updated + 43200) > ?', (self.timestamp,))
        tupsByID = cur.fetchall()
        brl_rateByID = {str(tup[0]) : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_statcast SET brl_rate = ? WHERE pitch_id = ?', ([(brl_rateByID.get(id, None), id) for id in arsenalIDs]))

        #Get hard hits by pitch id
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, COUNT(*) FROM trackman WHERE call_id = 4 AND exit_velocity >= 95 AND (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, COUNT(*) FROM trackman WHERE call_id = 4 AND exit_velocity >= 95 AND (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        hhByID = dict(cur.fetchall())
        cur.executemany('UPDATE arsenal_stats_statcast SET hh = ? WHERE pitch_id = ?', ([(hhByID.get(id, 0), id) for id in arsenalIDs]))
        
        #Get hard hit rate by pitch id
        cur.execute('SELECT pitch_id, bbe, hh FROM arsenal_stats_statcast WHERE (last_updated + 43200) > ?', (self.timestamp,))
        tupsByID = cur.fetchall()
        hh_rateByID = {str(tup[0]) : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_statcast SET hh_rate = ? WHERE pitch_id = ?', ([(hh_rateByID.get(id, None), id) for id in arsenalIDs]))
        
        #Get babip by pitch id
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, COUNT(*), COUNT(CASE WHEN result_id <= 4 THEN 1 END) FROM trackman WHERE call_id = 4 AND (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, COUNT(*), COUNT(CASE WHEN result_id <= 4 THEN 1 END) FROM trackman WHERE call_id = 4 AND (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        tupsByID = cur.fetchall()
        babipByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_statcast SET babip = ? WHERE pitch_id = ?', ([(babipByID.get(id, None), id) for id in arsenalIDs]))
        
        # Arsenal Batted Ball

        #Clear batted ball table of recent players if they are already listed
        cur.executemany('DELETE FROM arsenal_stats_batted_ball WHERE pitch_id = ?', [(id,) for id in arsenalIDs])
        
        #Insert every pitch id into the table that played in recently uploaded games
        cur.executemany('INSERT INTO arsenal_stats_batted_ball (pitch_id, pitcher_id, league_id, division_id, team_id, year, type_id, last_updated) VALUES (?,?,?,?,?,?,?,?)', (tupIDs))

        #Insert every batted ball event by pitch id
        cur.executemany('UPDATE arsenal_stats_batted_ball SET bbe = ? WHERE  pitch_id = ?', ([(bbeByID.get(id, 0), id) for id in arsenalIDs]))
        
        #Get every ground ball rate by pitch id
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, COUNT(*), COUNT(CASE WHEN launch_angle <= 10 THEN 1 END) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, COUNT(*), COUNT(CASE WHEN launch_angle <= 10 THEN 1 END) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        tupsByID = cur.fetchall()
        gb_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_batted_ball SET gb_rate = ? WHERE pitch_id = ?', ([(gb_rateByID.get(id, None), id) for id in arsenalIDs]))

        #Get every fly ball rate by pitch id 
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, COUNT(*), COUNT(CASE WHEN launch_angle > 25 AND launch_angle <=50 THEN 1 END) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, COUNT(*), COUNT(CASE WHEN launch_angle > 25 AND launch_angle <=50 THEN 1 END) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        tupsByID = cur.fetchall()
        fb_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_batted_ball SET fb_rate = ? WHERE pitch_id = ?', ([(fb_rateByID.get(id, None), id) for id in arsenalIDs]))

        #Get every line drive rate by pitch id
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, COUNT(*), COUNT(CASE WHEN launch_angle > 10 AND launch_angle <=25 THEN 1 END) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, COUNT(*), COUNT(CASE WHEN launch_angle > 10 AND launch_angle <=25 THEN 1 END) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        tupsByID = cur.fetchall()
        ld_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_batted_ball SET ld_rate = ? WHERE pitch_id = ?', ([(ld_rateByID.get(id, None), id) for id in arsenalIDs]))

        #Get every infield fly ball rate rate by pitch id
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, COUNT(*), COUNT(CASE WHEN launch_angle > 50 THEN 1 END) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, COUNT(*), COUNT(CASE WHEN launch_angle > 50 THEN 1 END) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        tupsByID = cur.fetchall()
        iffb_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_batted_ball SET iffb_rate = ? WHERE pitch_id = ?', ([(iffb_rateByID.get(id, None), id) for id in arsenalIDs]))

        #Get every home run per fly ball rate by pitch id
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, COUNT(CASE WHEN launch_angle > 25 AND launch_angle <=50 THEN 1 END), COUNT(CASE WHEN launch_angle > 25 AND launch_angle <=50 AND result_id = 4 THEN 1 END) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, COUNT(CASE WHEN launch_angle > 25 AND launch_angle <=50 THEN 1 END), COUNT(CASE WHEN launch_angle > 25 AND launch_angle <=50 AND result_id = 4 THEN 1 END) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        tupsByID = cur.fetchall()
        hr_fbByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_batted_ball SET hr_fb = ? WHERE pitch_id = ?', ([(hr_fbByID.get(id, None), id) for id in arsenalIDs]))

        #Get every strike rate rate by pitch id
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, COUNT(*), COUNT(CASE WHEN call_id <= 4 THEN 1 END) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, COUNT(*), COUNT(CASE WHEN call_id <= 4 THEN 1 END) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        tupsByID = cur.fetchall()
        strike_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_batted_ball SET strike_rate = ? WHERE pitch_id = ?', ([(strike_rateByID.get(id, None), id) for id in arsenalIDs]))

        #Get every ball rate rate by pitch id
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, COUNT(*), COUNT(CASE WHEN call_id >= 5 AND call_id <= 7 THEN 1 END) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, COUNT(*), COUNT(CASE WHEN call_id >= 5 AND call_id <= 7 THEN 1 END) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        tupsByID = cur.fetchall()
        ball_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_batted_ball SET ball_rate = ? WHERE pitch_id = ?', ([(ball_rateByID.get(id, None), id) for id in arsenalIDs]))
 
        #Get every pull rate by pitch id
        cur.execute('''SELECT "1" || trackman.pitcher_id || trackman.tagged_type_id as tagged_pitch_id, COUNT(*),
        COUNT(CASE WHEN (trackman.hit_bearing < -15 AND batters.batter_side_id = 1) OR (trackman.hit_bearing > 15 AND batters.batter_side_id = 2) OR
        (trackman.hit_bearing > 15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 1) OR
        (trackman.hit_bearing < -15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 2) THEN 1 END)
        FROM trackman
        LEFT JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
        LEFT JOIN batters ON trackman.batter_id = batters.batter_id
        WHERE trackman.call_id = 4 AND trackman.exit_velocity > 0 AND trackman.hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || trackman.pitcher_id || trackman.auto_type_id as auto_pitch_id, COUNT(*),
        COUNT(CASE WHEN (trackman.hit_bearing < -15 AND batters.batter_side_id = 1) OR (trackman.hit_bearing > 15 AND batters.batter_side_id = 2) OR
        (trackman.hit_bearing > 15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 1) OR
        (trackman.hit_bearing < -15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 2) THEN 1 END)
        FROM trackman
        LEFT JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
        LEFT JOIN batters ON trackman.batter_id = batters.batter_id
        WHERE trackman.call_id = 4 AND trackman.exit_velocity > 0 AND trackman.hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        tupsByID = cur.fetchall()
        pull_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_batted_ball SET pull_rate = ? WHERE pitch_id = ?', ([(pull_rateByID.get(id, None), id) for id in arsenalIDs]))

        #Get every center rate by pitch id
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, COUNT(*),
        COUNT(CASE WHEN hit_bearing >= -15 AND hit_bearing <= 15 THEN 1 END)
        FROM trackman
        WHERE call_id = 4 AND exit_velocity > 0 AND hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, COUNT(*),
        COUNT(CASE WHEN hit_bearing >= -15 AND hit_bearing <= 15 THEN 1 END)
        FROM trackman
        WHERE call_id = 4 AND exit_velocity > 0 AND hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        tupsByID = cur.fetchall()
        cent_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_batted_ball SET cent_rate = ? WHERE pitch_id = ?', ([(cent_rateByID.get(id, None), id) for id in arsenalIDs]))

        #Get every oppo rate by pitch id
        cur.execute('''SELECT "1" || trackman.pitcher_id || trackman.tagged_type_id as tagged_pitch_id, COUNT(*),
        COUNT(CASE WHEN (trackman.hit_bearing > 15 AND batters.batter_side_id = 1) OR (trackman.hit_bearing < -15 AND batters.batter_side_id = 2) OR
        (trackman.hit_bearing < -15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 1) OR
        (trackman.hit_bearing > 15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 2) THEN 1 END)
        FROM trackman
        LEFT JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
        LEFT JOIN batters ON trackman.batter_id = batters.batter_id
        WHERE trackman.call_id = 4 AND trackman.exit_velocity > 0 AND trackman.hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || trackman.pitcher_id || trackman.auto_type_id as auto_pitch_id, COUNT(*),
        COUNT(CASE WHEN (trackman.hit_bearing > 15 AND batters.batter_side_id = 1) OR (trackman.hit_bearing < -15 AND batters.batter_side_id = 2) OR
        (trackman.hit_bearing < -15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 1) OR
        (trackman.hit_bearing > 15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 2) THEN 1 END)
        FROM trackman
        LEFT JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
        LEFT JOIN batters ON trackman.batter_id = batters.batter_id
        WHERE trackman.call_id = 4 AND trackman.exit_velocity > 0 AND trackman.hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        tupsByID = cur.fetchall()
        oppo_rateByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_batted_ball SET oppo_rate = ? WHERE pitch_id = ?', ([(oppo_rateByID.get(id, None), id) for id in arsenalIDs]))

        #Get every pull per gb rate by pitch id 
        cur.execute('''SELECT "1" || trackman.pitcher_id || trackman.tagged_type_id as tagged_pitch_id, COUNT(*),
        COUNT(CASE WHEN (trackman.hit_bearing < -15 AND batters.batter_side_id = 1) OR (trackman.hit_bearing > 15 AND batters.batter_side_id = 2) OR
        (trackman.hit_bearing > 15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 1) OR
        (trackman.hit_bearing < -15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 2) THEN 1 END)
        FROM trackman
        LEFT JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
        LEFT JOIN batters ON trackman.batter_id = batters.batter_id
        WHERE trackman.call_id = 4 AND trackman.exit_velocity > 0  AND trackman.launch_angle <= 10 AND trackman.hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || trackman.pitcher_id || trackman.auto_type_id as auto_pitch_id, COUNT(*),
        COUNT(CASE WHEN (trackman.hit_bearing < -15 AND batters.batter_side_id = 1) OR (trackman.hit_bearing > 15 AND batters.batter_side_id = 2) OR
        (trackman.hit_bearing > 15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 1) OR
        (trackman.hit_bearing < -15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 2) THEN 1 END)
        FROM trackman
        LEFT JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
        LEFT JOIN batters ON trackman.batter_id = batters.batter_id
        WHERE trackman.call_id = 4 AND trackman.exit_velocity > 0  AND trackman.launch_angle <= 10 AND trackman.hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        tupsByID = cur.fetchall()
        pull_gbByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_batted_ball SET pull_gb_rate = ? WHERE pitch_id = ?', ([(pull_gbByID.get(id, None), id) for id in arsenalIDs]))
 
        #Get every center per gb rate by pitch id
        cur.execute('''SELECT "1" || trackman.pitcher_id || trackman.tagged_type_id as tagged_pitch_id, COUNT(*),
        COUNT(CASE WHEN hit_bearing >= -15 AND hit_bearing <= 15 THEN 1 END)
        FROM trackman
        WHERE call_id = 4 AND exit_velocity > 0 AND launch_angle <= 10 AND hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || trackman.pitcher_id || trackman.auto_type_id as auto_pitch_id, COUNT(*),
        COUNT(CASE WHEN hit_bearing >= -15 AND hit_bearing <= 15 THEN 1 END)
        FROM trackman
        WHERE call_id = 4 AND exit_velocity > 0 AND launch_angle <= 10 AND hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        tupsByID = cur.fetchall()
        cent_gbByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_batted_ball SET cent_gb_rate = ? WHERE pitch_id = ?', ([(cent_gbByID.get(id, None), id) for id in arsenalIDs]))

        #Get every oppo per gb rate by pitch id
        cur.execute('''SELECT "1" || trackman.pitcher_id || trackman.tagged_type_id as tagged_pitch_id, COUNT(*),
        COUNT(CASE WHEN (trackman.hit_bearing > 15 AND batters.batter_side_id = 1) OR (trackman.hit_bearing < -15 AND batters.batter_side_id = 2) OR
        (trackman.hit_bearing < -15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 1) OR
        (trackman.hit_bearing > 15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 2) THEN 1 END)
        FROM trackman
        LEFT JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
        LEFT JOIN batters ON trackman.batter_id = batters.batter_id
        WHERE trackman.call_id = 4 AND trackman.exit_velocity > 0 AND trackman.launch_angle <= 10 AND trackman.hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || trackman.pitcher_id || trackman.auto_type_id as auto_pitch_id, COUNT(*),
        COUNT(CASE WHEN (trackman.hit_bearing > 15 AND batters.batter_side_id = 1) OR (trackman.hit_bearing < -15 AND batters.batter_side_id = 2) OR
        (trackman.hit_bearing < -15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 1) OR
        (trackman.hit_bearing > 15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 2) THEN 1 END)
        FROM trackman
        LEFT JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
        LEFT JOIN batters ON trackman.batter_id = batters.batter_id
        WHERE trackman.call_id = 4 AND trackman.exit_velocity > 0 AND trackman.launch_angle <= 10 AND trackman.hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        tupsByID = cur.fetchall()
        oppo_gbByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_batted_ball SET oppo_gb_rate = ? WHERE pitch_id = ?', ([(oppo_gbByID.get(id, None), id) for id in arsenalIDs]))

        #Get every pull per fb rate by pitch id (use lower, lower bound for launch angle for purpose of stat--to be used for outfield positioning)
        cur.execute('''SELECT "1" || trackman.pitcher_id || trackman.tagged_type_id as tagged_pitch_id, COUNT(*),
        COUNT(CASE WHEN (trackman.hit_bearing < -15 AND batters.batter_side_id = 1) OR (trackman.hit_bearing > 15 AND batters.batter_side_id = 2) OR
        (trackman.hit_bearing > 15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 1) OR
        (trackman.hit_bearing < -15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 2) THEN 1 END)
        FROM trackman
        LEFT JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
        LEFT JOIN batters ON trackman.batter_id = batters.batter_id
        WHERE trackman.call_id = 4 AND trackman.exit_velocity > 0  AND trackman.launch_angle > 20 AND trackman.launch_angle <=50 AND trackman.hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || trackman.pitcher_id || trackman.auto_type_id as auto_pitch_id, COUNT(*),
        COUNT(CASE WHEN (trackman.hit_bearing < -15 AND batters.batter_side_id = 1) OR (trackman.hit_bearing > 15 AND batters.batter_side_id = 2) OR
        (trackman.hit_bearing > 15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 1) OR
        (trackman.hit_bearing < -15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 2) THEN 1 END)
        FROM trackman
        LEFT JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
        LEFT JOIN batters ON trackman.batter_id = batters.batter_id
        WHERE trackman.call_id = 4 AND trackman.exit_velocity > 0  AND trackman.launch_angle > 20 AND trackman.launch_angle <=50 AND trackman.hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        tupsByID = cur.fetchall()
        pull_fbByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_batted_ball SET pull_fb_rate = ? WHERE pitch_id = ?', ([(pull_fbByID.get(id, None), id) for id in arsenalIDs]))

        #Get every center per fb rate by pitch id (use lower, lower bound for launch angle for purpose of stat--to be used for outfield positioning)
        cur.execute('''SELECT "1" || trackman.pitcher_id || trackman.tagged_type_id as tagged_pitch_id, COUNT(*),
        COUNT(CASE WHEN hit_bearing >= -15 AND hit_bearing <= 15 THEN 1 END)
        FROM trackman
        WHERE call_id = 4 AND exit_velocity > 0 AND launch_angle > 20 AND launch_angle <=50 AND hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || trackman.pitcher_id || trackman.auto_type_id as auto_pitch_id, COUNT(*),
        COUNT(CASE WHEN hit_bearing >= -15 AND hit_bearing <= 15 THEN 1 END)
        FROM trackman
        WHERE call_id = 4 AND exit_velocity > 0 AND launch_angle > 20 AND launch_angle <=50 AND hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        tupsByID = cur.fetchall()
        cent_fbByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_batted_ball SET cent_fb_rate = ? WHERE pitch_id = ?', ([(cent_fbByID.get(id, None), id) for id in arsenalIDs]))

        #Get every oppo per fb rate by pitch id (use lower, lower bound for launch angle for purpose of stat--to be used for outfield positioning)
        cur.execute('''SELECT "1" || trackman.pitcher_id || trackman.tagged_type_id as tagged_pitch_id, COUNT(*),
        COUNT(CASE WHEN (trackman.hit_bearing > 15 AND batters.batter_side_id = 1) OR (trackman.hit_bearing < -15 AND batters.batter_side_id = 2) OR
        (trackman.hit_bearing < -15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 1) OR
        (trackman.hit_bearing > 15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 2) THEN 1 END)
        FROM trackman
        LEFT JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
        LEFT JOIN batters ON trackman.batter_id = batters.batter_id
        WHERE trackman.call_id = 4 AND trackman.exit_velocity > 0 AND trackman.launch_angle > 20 AND trackman.launch_angle <=50 AND trackman.hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || trackman.pitcher_id || trackman.auto_type_id as auto_pitch_id, COUNT(*),
        COUNT(CASE WHEN (trackman.hit_bearing > 15 AND batters.batter_side_id = 1) OR (trackman.hit_bearing < -15 AND batters.batter_side_id = 2) OR
        (trackman.hit_bearing < -15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 1) OR
        (trackman.hit_bearing > 15 AND batters.batter_side_id = 3 AND pitchers.pitcher_side_id = 2) THEN 1 END)
        FROM trackman
        LEFT JOIN pitchers ON trackman.pitcher_id = pitchers.pitcher_id
        LEFT JOIN batters ON trackman.batter_id = batters.batter_id
        WHERE trackman.call_id = 4 AND trackman.exit_velocity > 0 AND trackman.launch_angle > 20 AND trackman.launch_angle <=50 AND trackman.hit_bearing IS NOT NULL AND (trackman.upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        tupsByID = cur.fetchall()
        oppo_fbByID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_batted_ball SET oppo_fb_rate = ? WHERE pitch_id = ?', ([(oppo_fbByID.get(id, None), id) for id in arsenalIDs]))

        #Get soft con rate by pitch id
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, COUNT(*), COUNT(CASE WHEN exit_velocity < 70 THEN 1 END) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, COUNT(*), COUNT(CASE WHEN exit_velocity < 70 THEN 1 END) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        tupsByID = cur.fetchall()
        soft_conbyID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_batted_ball SET soft_con_rate = ? WHERE pitch_id = ?', ([(soft_conbyID.get(id, None), id) for id in arsenalIDs]))
 
        #Get med con rate by pitch id
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, COUNT(*), COUNT(CASE WHEN exit_velocity >= 70 AND exit_velocity < 95 THEN 1 END) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, COUNT(*), COUNT(CASE WHEN exit_velocity >= 70 AND exit_velocity < 95 THEN 1 END) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        tupsByID = cur.fetchall()
        med_conbyID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_batted_ball SET med_con_rate = ? WHERE pitch_id = ?', ([(med_conbyID.get(id, None), id) for id in arsenalIDs]))
  
        #Get hard con rate by pitch id
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, COUNT(*), COUNT(CASE WHEN exit_velocity >= 95 THEN 1 END) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, COUNT(*), COUNT(CASE WHEN exit_velocity >= 95 THEN 1 END) FROM trackman WHERE call_id = 4 AND exit_velocity > 0 AND (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        tupsByID = cur.fetchall()
        hard_conbyID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_batted_ball SET hard_con_rate = ? WHERE pitch_id = ?', ([(hard_conbyID.get(id, None), id) for id in arsenalIDs]))
        
        # Discipline 

        #Clear discipline table of recent players if they are already listed
        cur.executemany('DELETE FROM arsenal_stats_discipline WHERE pitch_id = ?', [(id,) for id in arsenalIDs])
        
        #Insert every pitch id into the table that played in recently uploaded games
        cur.executemany('INSERT INTO arsenal_stats_discipline (pitch_id, pitcher_id, league_id, division_id, team_id, year, type_id, last_updated) VALUES (?,?,?,?,?,?,?,?)', (tupIDs))
        
        #Get outside of zone swing rate by pitch id
        #Get balls based on universal strike zone
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, COUNT(*), COUNT(CASE WHEN call_id = 2 OR call_id = 3 OR call_id = 4 THEN 1 END) FROM trackman WHERE location_side < -0.7508 OR location_side > 0.7508 OR location_height < 1.5942 OR location_height > 3.6033 AND (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, COUNT(*), COUNT(CASE WHEN call_id = 2 OR call_id = 3 OR call_id = 4 THEN 1 END) FROM trackman WHERE location_side < -0.7508 OR location_side > 0.7508 OR location_height < 1.5942 OR location_height > 3.6033 AND (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        tupsByID = cur.fetchall()
        o_swingbyID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_discipline SET o_swing = ? WHERE pitch_id = ?', ([(o_swingbyID.get(id, None), id) for id in arsenalIDs]))

        #Get inside of zone swing rate by pitch id
        #Get strikes based on universal strike zone
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, COUNT(*), COUNT(CASE WHEN call_id = 2 OR call_id = 3 OR call_id = 4 THEN 1 END) FROM trackman WHERE location_side >= -0.7508 AND location_side <= 0.7508 AND location_height >= 1.5942 AND location_height <= 3.6033 AND (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, COUNT(*), COUNT(CASE WHEN call_id = 2 OR call_id = 3 OR call_id = 4 THEN 1 END) FROM trackman WHERE location_side >= -0.7508 AND location_side <= 0.7508 AND location_height >= 1.5942 AND location_height <= 3.6033 AND (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        tupsByID = cur.fetchall()
        z_swingbyID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_discipline SET z_swing = ? WHERE pitch_id = ?', ([(z_swingbyID.get(id, None), id) for id in arsenalIDs]))
        
        #Get swing rate by pitch id
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, COUNT(*), COUNT(CASE WHEN call_id = 2 OR call_id = 3 OR call_id = 4 THEN 1 END) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, COUNT(*), COUNT(CASE WHEN call_id = 2 OR call_id = 3 OR call_id = 4 THEN 1 END) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        tupsByID = cur.fetchall()
        swing_ratebyID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_discipline SET swing_rate = ? WHERE pitch_id = ?', ([(swing_ratebyID.get(id, None), id) for id in arsenalIDs]))
        
        #Get outside of zone contact rate by pitch id
        #Get balls based on universal strike zone
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, COUNT(CASE WHEN call_id = 2 OR call_id = 3 OR call_id = 4 THEN 1 END), COUNT(CASE WHEN call_id = 3 OR call_id = 4 THEN 1 END) FROM trackman WHERE location_side < -0.7508 OR location_side > 0.7508 OR location_height < 1.5942 OR location_height > 3.6033 AND (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, COUNT(CASE WHEN call_id = 2 OR call_id = 3 OR call_id = 4 THEN 1 END), COUNT(CASE WHEN call_id = 3 OR call_id = 4 THEN 1 END) FROM trackman WHERE location_side < -0.7508 OR location_side > 0.7508 OR location_height < 1.5942 OR location_height > 3.6033 AND (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        tupsByID = cur.fetchall()
        o_contactbyID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_discipline SET o_contact = ? WHERE pitch_id = ?', ([(o_contactbyID.get(id, None), id) for id in arsenalIDs]))
        
        #Get inside of zone contact rate by pitch id
        #Get strikes based on universal strike zone
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, COUNT(CASE WHEN call_id = 2 OR call_id = 3 OR call_id = 4 THEN 1 END), COUNT(CASE WHEN call_id = 3 OR call_id = 4 THEN 1 END) FROM trackman WHERE location_side >= -0.7508 AND location_side <= 0.7508 AND location_height >= 1.5942 AND location_height <= 3.6033 AND (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, COUNT(CASE WHEN call_id = 2 OR call_id = 3 OR call_id = 4 THEN 1 END), COUNT(CASE WHEN call_id = 3 OR call_id = 4 THEN 1 END) FROM trackman WHERE location_side >= -0.7508 AND location_side <= 0.7508 AND location_height >= 1.5942 AND location_height <= 3.6033 AND (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        tupsByID = cur.fetchall()
        z_contactbyID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_discipline SET z_contact = ? WHERE pitch_id = ?', ([(z_contactbyID.get(id, None), id) for id in arsenalIDs]))
        
        #Get contact rate by pitch id
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, COUNT(CASE WHEN call_id = 2 OR call_id = 3 OR call_id = 4 THEN 1 END), COUNT(CASE WHEN call_id = 3 OR call_id = 4 THEN 1 END) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, COUNT(CASE WHEN call_id = 2 OR call_id = 3 OR call_id = 4 THEN 1 END), COUNT(CASE WHEN call_id = 3 OR call_id = 4 THEN 1 END) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        tupsByID = cur.fetchall()
        contact_ratebyID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_discipline SET contact_rate = ? WHERE pitch_id = ?', ([(contact_ratebyID.get(id, None), id) for id in arsenalIDs]))

        #Get inside of zone rate by pitch id
        #Get strikes based on universal strike zone
        cur.execute('''SELECT "1" || pitcher_id || tagged_type_id as tagged_pitch_id, COUNT(*), COUNT(CASE WHEN location_side >= -0.7508 AND location_side <= 0.7508 AND location_height >= 1.5942 AND location_height <= 3.6033 THEN 1 END) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY tagged_pitch_id
        UNION SELECT "2" || pitcher_id || auto_type_id as auto_pitch_id, COUNT(*), COUNT(CASE WHEN location_side >= -0.7508 AND location_side <= 0.7508 AND location_height >= 1.5942 AND location_height <= 3.6033 THEN 1 END) FROM trackman WHERE (upload_timestamp + 43200) > ? GROUP BY auto_pitch_id''', (self.timestamp,self.timestamp))
        tupsByID = cur.fetchall()
        zone_ratebyID = {tup[0] : tup[2] / tup[1] for tup in tupsByID if tup[1]}
        cur.executemany('UPDATE arsenal_stats_discipline SET zone_rate = ? WHERE pitch_id = ?', ([(zone_ratebyID.get(id, None), id) for id in arsenalIDs]))

        #Do these on the dashboard side, should be a simple query (player / league)
        # Arsenal Standard Plus Strike%+, H/TBF+, HR/TBF+
        # Arsenal Info Plus (compared to pitch type) velo+, maxVelo+, spinrate+, vert movement+, induced vert+, horz movement+
        # Arsenal Statcast Plus EV+, maxEV+, Barrel%+, HH%+, BABIP+, xBABIP+,
        # Arsenal Batted Ball Plus GB/FB, LD%, GB%, FB%, IFFB%, HR/FB, RS, RS/9, PULL%, CENT%, OPPO%, PULLGB%, CENTGB%, OPPOGB%, PULLOFFB%, CENTOFFB%, OPPOOFFB%, SOFT%, MED%, HARD%
        # Arsenal Discipline Plus O-Swing%, Z-Swing%, Swing%, O-Contact%, Z-Contact%, Contact%, Zone%

        conn.commit()
        clear_output(wait=True)
        print('Done!')
        cur.close()